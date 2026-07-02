"""Render report rows to a styled .xlsx workbook (openpyxl) and return bytes."""
import io
import datetime as _dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BRAND = "1F6F8B"          # BI - Reporting brand
_HEADER_FILL = PatternFill("solid", fgColor=_BRAND)
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_TITLE_FONT = Font(bold=True, size=14, color=_BRAND, name="Calibri")
_META_FONT = Font(size=9, color="666666", italic=True)
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def build_xlsx(columns, rows, title="Report", meta=None, sheet_name="Report"):
    wb = Workbook()
    _fill_sheet(wb.active, columns, rows, title, meta, sheet_name)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_xlsx_multi(sections, title="Report", meta=None):
    """Multi-section workbook: one styled sheet per section.

    sections = [{key, title, columns, rows}] (a MULTI definition's output).
    Key-value sections render like any table (their single row spreads down
    naturally in two-column form is not needed for Excel consumers).
    """
    wb = Workbook()
    first = True
    for i, s in enumerate(sections, start=1):
        ws = wb.active if first else wb.create_sheet()
        first = False
        sheet = f"{i}. {s.get('title') or s.get('key') or 'Section'}"
        _fill_sheet(ws, s["columns"], s["rows"], s.get("title") or title, meta, sheet)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BAD_SHEET = str.maketrans({c: " " for c in "[]:*?/\\"})


def _fill_sheet(ws, columns, rows, title, meta, sheet_name):
    ws.title = (str(sheet_name or "Report").translate(_BAD_SHEET))[:31]

    ncols = max(1, len(columns))
    # title + meta band
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    if meta:
        c = ws.cell(row=2, column=1, value=meta)
        c.font = _META_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    head_row = 4
    for j, col in enumerate(columns, start=1):
        c = ws.cell(row=head_row, column=j, value=str(col).replace("_", " ").title())
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER

    widths = [len(str(col)) + 2 for col in columns]
    for i, row in enumerate(rows, start=head_row + 1):
        for j, val in enumerate(row, start=1):
            if isinstance(val, _dt.datetime):
                val = val.strftime("%Y-%m-%d %H:%M")
            elif isinstance(val, _dt.date):
                val = val.strftime("%Y-%m-%d")
            c = ws.cell(row=i, column=j, value=val)
            c.border = _BORDER
            if _is_number(val):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            widths[j - 1] = max(widths[j - 1], len(str(val)) + 2)

    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(10, w))
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
