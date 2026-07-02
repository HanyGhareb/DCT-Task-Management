"""Unit tests for the render layer (no DB / no browser needed).

Run from reporting/runner/:  python -m pytest tests -q
"""
import os
import sys
import datetime as dt

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import render_pdf      # noqa: E402
import render_xlsx     # noqa: E402

COLUMNS = ["period_name", "sector_name", "budget_ytd", "gl_actual_ytd", "period_date"]
ROWS = [
    ("05-2026", "Infrastructure", 1234567.5, 1000000.0, dt.date(2026, 5, 1)),
    ("05-2026", "Culture & Arts", 250000.0, 312500.25, dt.date(2026, 5, 1)),
]
CTX = {
    "report_code": "GL_BUDGET_ACTUAL", "report_name": "Budget vs Actual (YTD)",
    "period": "05-2026", "generated_at": "2026-06-30 10:23 PM", "row_count": len(ROWS),
    "headers": [c.replace("_", " ").title() for c in COLUMNS],
    "columns": COLUMNS, "rows": ROWS, "meta": "Generated 2026-06-30 | 2 rows", "brand": "#1F6F8B",
}


def test_html_has_title_and_formatted_numbers():
    html = render_pdf.render_html(CTX)
    assert "Budget vs Actual (YTD)" in html
    assert "1,234,567.50" in html          # float formatted with thousands + 2dp
    assert "Sector Name" in html            # titleized header
    assert html.count("<tr>") >= 3          # header + 2 data rows


def test_html_empty_state():
    html = render_pdf.render_html(dict(CTX, rows=[], row_count=0))
    assert "No data" in html


def test_xlsx_is_a_zip_workbook():
    data = render_xlsx.build_xlsx(COLUMNS, ROWS, title="Budget vs Actual", meta="2 rows")
    assert data[:2] == b"PK"                # xlsx is a zip container
    assert len(data) > 2000


def test_fmt_filter():
    assert render_pdf._fmt(1234567) == "1,234,567"
    assert render_pdf._fmt(None) == ""
    assert render_pdf._fmt(3.5) == "3.50"

# ── MULTI-section (BUDGET_UTIL_SECTOR-style) ─────────────────────────────────

SECTIONS = [
    {"key": "intro", "title": "Sector Overview", "layout": "kv",
     "columns": ["sector_code", "sector_name", "budget", "actual_ap", "fund_available"],
     "rows": [("TOURISM", "Tourism", 1000000.0, 20000.0, 450000.0)],
     "headers": ["Sector Code", "Sector Name", "Budget", "Actual Ap", "Fund Available"],
     "totals": None},
    {"key": "utilization", "title": "Budget Utilization", "layout": "table",
     "columns": ["department", "cost_centre", "project_number", "project_name",
                 "task_number", "gl_account", "chapter", "program", "expenditure_type",
                 "budget", "actual_ap", "actual_grn", "commitment_pr", "obligation_po",
                 "fund_available"],
     "rows": [("Dept A", "4515000", "P1", "Project One", "T1", "GL1", "Ch 2", "Prog",
               "Costs", 1000000.0, 20000.0, 500000.0, 20000.0, 10000.0, 450000.0)],
     "headers": ["Department", "Cost Centre", "Project Number", "Project Name",
                 "Task Number", "Gl Account", "Chapter", "Program", "Expenditure Type",
                 "Budget", "Actual Ap", "Actual Grn", "Commitment Pr", "Obligation Po",
                 "Fund Available"],
     "totals": [None] * 9 + [1000000.0, 20000.0, 500000.0, 20000.0, 10000.0, 450000.0]},
    {"key": "open_po", "title": "Open Purchase Orders", "layout": "table",
     "columns": ["po_number", "open_aed"], "rows": [],
     "headers": ["Po Number", "Open Aed"], "totals": [None, None]},
]
MULTI_CTX = {
    "report_code": "BUDGET_UTIL_SECTOR",
    "report_name": "Budget Utilization by Sector (Executive)",
    "period": None, "params": {"year": 2026, "sector": "Tourism"},
    "generated_at": "2026-07-02 06:55 PM", "row_count": 2, "requested_by": "test",
    "headers": [], "columns": [], "rows": [], "sections": SECTIONS,
    "meta": "Generated 2026-07-02 | 2 lines | year 2026 | sector Tourism",
    "brand": "#1F6F8B", "landscape": True,
}


def test_multi_template_sections_and_grouped_header():
    html = render_pdf.render_html(MULTI_CTX, template_name="budget_util_sector.html.j2")
    assert "Sector Overview" in html and "Budget Utilization" in html
    assert "Encumbrances" in html and ">Actual<" in html   # grouped 2-row header
    assert "A4 landscape" in html
    assert "1,000,000.00" in html                          # totals row formatted
    assert "No data for the selected criteria." in html    # empty open_po section


def test_multi_xlsx_one_sheet_per_section():
    data = render_xlsx.build_xlsx_multi(SECTIONS, title="Budget Utilization by Sector")
    assert data[:2] == b"PK"
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert len(wb.sheetnames) == 3
    assert wb.sheetnames[0].startswith("1. ")


def test_multi_totals_money_columns_only():
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    import runner
    cols = ["po_number", "po_line", "line_aed", "open_aed"]
    rows = [("PO1", 1, 100.0, 40.0), ("PO2", 2, 50.0, 50.0)]
    totals = runner._totals(cols, rows)
    assert totals == [None, None, 150.0, 90.0]             # line numbers NOT summed


def test_fetch_multi_requires_params():
    import datasource
    import json
    import pytest
    spec = json.dumps({"required": ["year", "sector"], "sections": []})
    with pytest.raises(ValueError, match="sector"):
        datasource.fetch_multi(None, spec, {"year": 2026})
