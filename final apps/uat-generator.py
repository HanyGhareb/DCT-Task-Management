# -*- coding: utf-8 -*-
"""Generate UAT workbooks for the 6 i-Finance JET apps.

One .xlsx per app under final apps/<App>/UAT/, named
UAT_<APP>_dd-Mon-YYYY-NN.xlsx (NN auto-increments — re-runs never overwrite):
  - Instructions sheet (how to test, legend, seeded data, live summary)
  - One sheet per logical functional area
  - Columns: Test ID | Function | Test Scenario | Steps |
             Sample Data / Records | Expected Result |
             Status (dropdown) | Tested By | Test Date | Comments / Defect Ref
  - Pass/Fail/Blocked/N/A dropdown + conditional colors, frozen header, autofilter
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ROOT = r"c:\claude\DCT-task-management\DCT-Task-Management\final apps"

BRANDS = {  # header color per app
    "Admin": "C74634", "PC": "2E7D32", "DT": "0572CE", "HR": "1A7F5A",
    "FL": "7C4DBE", "CC": "B0721E",
}

HEADERS = ["Test ID", "Function", "Test Scenario", "Steps", "Sample Data / Records",
           "Expected Result", "Status", "Tested By", "Test Date", "Comments / Defect Ref"]
WIDTHS  = [10, 20, 30, 46, 34, 42, 11, 14, 12, 32]
STATUS_COL = "G"   # Status column letter — keep in sync with HEADERS

thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# =============================================================================
# Test case data.  Per app: list of (area_sheet_name, prefix, [cases])
# case = (function, scenario, steps, expected)
# =============================================================================

ADMIN = [
("Login & Session", "LOG", [
 ("Login", "Valid login", "1. Open the app URL\n2. Enter a valid username + password (or use a quick-login button)\n3. Click Sign In", "Dashboard opens; top bar shows your name and avatar initials; side navigation matches your role"),
 ("Login", "Invalid password", "1. Enter a valid username with a wrong password\n2. Click Sign In", "Error message 'Invalid credentials or account inactive'; stays on login page"),
 ("Login", "Inactive account", "1. Login with a deactivated account (ask admin to deactivate a test user first)", "Login refused with the same generic error"),
 ("Session", "Refresh keeps session", "1. Login\n2. Press F5 / Ctrl+F5 on any page", "Still logged in; same page or dashboard reopens (no login prompt)"),
 ("Session", "Deep link without session", "1. Logout\n2. Open the app URL again", "Login page shown; no app content visible before login"),
 ("Logout", "Logout clears session", "1. Click your avatar (top-right) > Logout\n2. Press Back in the browser", "Returned to login; Back does not re-enter the app with data"),
 ("Login", "Arabic login page", "1. On the login page switch language to AR (ع)\n2. Login", "Login page is mirrored (RTL) with Arabic labels; after login the app stays in Arabic"),
]),
("Dashboard", "DSH", [
 ("Headline counts", "Stats cards show live numbers", "1. Open Home/Dashboard\n2. Compare 'Users' count with the Users list total", "Cards show non-zero sensible numbers; users count matches the Users page total"),
 ("Charts", "Both charts render", "1. Open Dashboard\n2. Wait for charts to load", "Approval cycle-time bar chart + 30-day activity line chart render with axes and legend; no empty boxes"),
 ("Charts", "Charts survive navigation", "1. Open Dashboard\n2. Navigate to Users, then back to Dashboard", "Charts render again correctly (not blank, not duplicated)"),
 ("Loading states", "Skeleton then data", "1. Hard-refresh (Ctrl+F5) on Dashboard", "Brief skeleton/placeholder shown, then real data; no flash of errors"),
]),
("My Workspace", "WSP", [
 ("My Profile", "Profile loads current data", "1. Go to My Workspace > My Profile", "Display names, email, phone, employee number, organisation are the CURRENT values from the server"),
 ("My Profile", "Update phone persists", "1. Change Phone\n2. Click Save Profile\n3. Hard-refresh (Ctrl+F5) and reopen My Profile", "Success banner on save; after refresh the NEW phone is still shown"),
 ("My Profile", "Update Arabic display name", "1. Edit Display Name (Arabic)\n2. Save\n3. Refresh and recheck", "Arabic name saved and persists; shown RTL in the field"),
 ("My Profile", "Employee number not lost", "1. Note your Employee Number\n2. Change phone and Save Profile\n3. Refresh", "Employee Number is unchanged after a profile save"),
 ("My Profile", "Upload photo", "1. Click Change Photo\n2. Pick a JPG/PNG (any size)\n3. Wait for upload", "File picker opens; avatar shows the photo; photo still there after Ctrl+F5"),
 ("My Profile", "Replace photo", "1. Upload a different photo over the existing one", "New photo replaces the old one without errors"),
 ("Change password", "Happy path", "1. In Change Password enter a new valid password twice\n2. Submit\n3. Logout and login with the NEW password\n4. Change it back", "Success message; old password no longer works; new one does"),
 ("Change password", "Validation", "1. Try mismatching passwords\n2. Try a password shorter than 8 characters", "Clear validation errors; nothing saved"),
 ("Notifications", "List + unread badge", "1. Open Notifications\n2. Compare the bell badge count with unread rows", "List loads; badge equals the number of unread notifications"),
 ("Notifications", "Mark as read", "1. Mark one notification as read", "Row state changes; bell badge count decreases by 1"),
 ("Pending approvals", "My queue", "1. Open Pending Approvals (use an approver account with pending items)", "Only items waiting for YOUR approval are listed with module, requester, date"),
 ("Pending approvals", "Approve / reject", "1. Approve one item\n2. Reject another with a comment", "Items leave the queue; requester gets a notification; status history updated"),
]),
("User Management", "USR", [
 ("Users list", "Server pagination", "1. Open Users\n2. Check the pager caption (e.g. '1-50 of N')\n3. Click next page", "Counts are correct; next page loads different rows; pager updates"),
 ("Users list", "Search", "1. Type part of a name/username in search\n2. Pause typing", "List filters after a short delay (server search); count updates; clearing search restores the full list"),
 ("Users list", "Status filter", "1. Filter by Active, then Inactive", "Only matching users shown each time"),
 ("Create user", "Happy path", "1. Click New User\n2. Fill username, names, email, org, role(s), password\n3. Save\n4. Logout and login as the new user", "User created and appears in the list; new user can login and sees role-appropriate navigation"),
 ("Create user", "Duplicate username", "1. Create a user with an existing username", "Clear error; no duplicate created"),
 ("Edit user", "Change roles", "1. Edit a test user\n2. Add/remove a role\n3. Save\n4. Login as that user", "Saved roles reflected; navigation/permissions change accordingly"),
 ("Edit user", "Edit fields", "1. Edit display names / email / phone / employee number / org\n2. Save and reopen", "All edited values persisted and shown"),
 ("Deactivate", "Deactivate + reactivate", "1. Deactivate a test user\n2. Try logging in as them\n3. Reactivate", "Deactivated user cannot login; after reactivation they can"),
]),
("Roles & Permissions", "ROL", [
 ("Roles list", "Vault list renders", "1. Open Roles", "Dark 'Vault' page lists roles with code, type, member count, active state"),
 ("Create role", "New role", "1. Create a role with code + names\n2. Save", "Role appears in the list; audit entry recorded"),
 ("Edit role", "Assign permissions", "1. Open a role\n2. Toggle several permissions on/off\n3. Save\n4. Reopen", "Saved toggles persist; member count unchanged"),
 ("Permission matrix", "Matrix view", "1. Open Permissions\n2. Locate a role column and toggle a cell\n3. Save", "Matrix shows roles x permissions; toggle animates; change persists after reload"),
 ("Permissions", "Effect on user", "1. Remove a permission from a role\n2. Login as a user with ONLY that role", "Corresponding function is hidden/refused for that user (restore afterwards)"),
]),
("Organisation", "ORG", [
 ("Hierarchy", "Tree renders", "1. Open Org Hierarchy", "Organisation tree shows with expand/collapse; correct parent-child structure"),
 ("Hierarchy", "Add unit", "1. Add a child org unit (EN + AR names)\n2. Save", "New unit appears under its parent; available in user org dropdown"),
 ("Hierarchy", "Edit / deactivate unit", "1. Rename a test unit\n2. Deactivate it", "Rename persists; deactivated unit flagged and not offered for new users"),
]),
("System Administration", "SYS", [
 ("Module registry", "List + toggle", "1. Open Module Registry\n2. Toggle a non-critical module inactive and back", "Modules (PC/DT/HR/FL/CC...) listed with status; toggle persists"),
 ("Approval templates", "View / edit template", "1. Open Approval Templates\n2. Open one template and review steps (sequence, role, escalation days)\n3. Edit a non-production template", "Steps listed in order; edits persist"),
 ("Approval monitor", "In-flight instances", "1. Open Approval Monitor\n2. Filter by module/status", "Live approval instances with current step; filters work"),
 ("Lookups", "Add lookup value", "1. Open Lookups\n2. Pick a category (e.g. a status set)\n3. Add a value (code + EN/AR names)\n4. Edit it, then disable it", "Value appears immediately; edits persist; disabled value stops appearing in module dropdowns"),
 ("System settings", "Edit a setting", "1. Open System Settings\n2. Edit a non-system setting value\n3. Save and refresh", "New value persists; system-flagged settings are read-only"),
 ("System settings", "Secrets masked", "1. Locate a key/password-type setting (e.g. API key)", "Value displays as ******** ; saving the mask back is refused (no secret overwrite)"),
 ("Theme setting", "Brand color from settings", "1. Edit THEME_BRAND_COLOR (e.g. #7B1FA2)\n2. Hard-refresh the app", "App chrome (buttons, accents) adopts the new color; restore the original afterwards"),
 ("Audit log", "Pagination + filter", "1. Open Audit Log\n2. Check pager caption\n3. Filter by action (e.g. LOGIN)\n4. Page through", "Server-paged list; filter narrows rows; newest entries first; your own recent actions appear"),
 ("Appearance", "Theme switch", "1. Open Appearance\n2. Switch theme and back", "Theme applies immediately and persists across refresh"),
]),
("Delegations & Announcements", "DLG", [
 ("My delegations", "Profile shows real data", "1. Open My Workspace > My Profile\n2. Scroll to My Delegations", "Real delegations from the server (not placeholders); outgoing show 'Delegating to', incoming show 'Acting for'"),
 ("Create delegation", "Delegate my authority", "1. In My Profile click + New Delegation\n2. Pick a user, end date (e.g. +7 days), reason\n3. Create", "Delegation appears ACTIVE in the list; the delegate now sees your pending approvals with an 'acting for' badge"),
 ("Acting for", "Delegate sees the queue", "1. Login as the delegate\n2. Open Pending Approvals (Admin) or the module's approvals page", "Items requiring the delegator's role are listed with an 'acting for <name>' badge; delegate can approve/reject"),
 ("Cancel delegation", "Stop the cover", "1. As the delegator (or SYS_ADMIN) cancel the ACTIVE delegation", "Status becomes CANCELLED; the delegate's queue no longer shows the delegated items"),
 ("Oversight", "System > Delegations", "1. Login as SYS_ADMIN\n2. Open System > Delegations\n3. Filter by status", "All delegations platform-wide with delegator, delegate, scope, window; expired ones flip to EXPIRED automatically"),
 ("Announcements admin", "Create an announcement", "1. Open System > Announcements\n2. + New Announcement: EN+AR titles, severity, audience\n3. Save", "Announcement listed; appears as a colored banner at the top of the targeted apps within a refresh"),
 ("Banner audience", "Module-targeted banner", "1. Create a WARNING announcement targeted at module CREDIT_CARDS\n2. Open the CC app, then the PC app", "Banner shows in CC only; PC does not show it; ALL announcements show everywhere"),
 ("Banner dismiss", "Dismiss persists for the session", "1. Click the X on a banner\n2. Navigate around\n3. Close the tab, reopen and login", "Dismissed banner stays hidden while the session lasts; reappears in a new session if still active"),
 ("Deactivate", "Pull an announcement", "1. In System > Announcements deactivate one", "It stops appearing in all apps (after refresh); stays in the admin list as inactive"),
]),
("Shell & Platform", "SHL", [
 ("Module switcher", "Switch to another module", "1. Click the module dropdown in the top bar\n2. Choose HR (or PC/DT)", "Target app opens WITHOUT a second login (shared session); switcher highlights the current module"),
 ("Module switcher", "FL + CC are live", "1. Open the module dropdown", "Freelancers (APP 203) and Credit Cards (APP 202) appear WITHOUT a 'soon' badge and open their apps"),
 ("Language", "EN to AR live switch", "1. Click the ع button in the top bar", "Entire page flips RTL instantly: top bar, side nav, content; Arabic font; numbers stay Latin digits"),
 ("Language", "AR persists across login", "1. Switch to AR\n2. Logout, close browser, reopen and login", "App opens in Arabic (server-stored preference); switch back to EN at the end"),
 ("Side navigation", "Collapse / expand", "1. Click the hamburger icon", "Side nav collapses/expands; content area adjusts"),
 ("Role-based nav", "Menu matches role", "1. Login as a non-admin user", "Admin-only groups (User Management, System) are hidden"),
 ("Error handling", "API failure feedback", "1. Disconnect network (or stop the proxy)\n2. Try loading a list", "Friendly error state with Retry + a toast; no frozen skeleton, no blank screen"),
]),
("Wave Enhancements", "WAV", [
 ("Bootstrap", "One-call /boot on login", "1. Open DevTools > Network\n2. Login", "A single GET /dct/boot fires right after login (instead of separate settings/prefs/unread calls); brand color and language apply immediately"),
 ("Landing page", "Role-based landing route", "1. As admin set system setting LANDING_MANAGER = profile\n2. Login as a MANAGER user (e.g. NASER.ALKHAJA)\n3. Restore LANDING_MANAGER = dashboard afterwards", "The manager lands on My Profile after login instead of the dashboard (when /boot answers within the 800ms cap; otherwise dashboard)"),
 ("Command palette", "Ctrl+K opens + navigates", "1. Press Ctrl+K anywhere\n2. Type 'role'\n3. Choose 'Roles' with arrow keys + Enter", "Vault-styled palette opens; Navigation group lists matching pages; Enter navigates to Roles and closes the palette"),
 ("Command palette", "Search users", "1. Press Ctrl+K\n2. Type at least 2 letters of a username (e.g. AYESHA)\n3. Pick the user result", "Users group shows matching users; selecting one opens that user's edit page"),
 ("Feature flags", "Kill switch disables palette", "1. System Settings: set FEATURE_COMMAND_PALETTE = N\n2. Reload the app and press Ctrl+K\n3. Restore Y afterwards", "With the flag off, Ctrl+K does nothing (no palette); after restoring Y + reload it works again"),
 ("Undo toast", "Deactivate user + Undo", "1. Users: click the deactivate (trash) icon on a test user\n2. Click Undo in the toast within 8 seconds", "User flips to Inactive immediately with a '<name> deactivated' toast carrying an Undo action; Undo restores Active and shows a reactivated toast"),
 ("Bulk actions", "Bulk deactivate", "1. Users: tick the checkboxes of two test users\n2. In the bulk bar click Deactivate, then Apply in the confirm dialog", "Bulk bar shows '2 selected'; after Apply a 'Applied to 2 user(s)' toast appears and both rows show Inactive"),
 ("Bulk actions", "Bulk activate restores", "1. Select the same two users\n2. Bulk bar > Activate > Apply", "Both rows show Active again"),
 ("Bulk actions", "Bulk assign role", "1. Select one test user\n2. Bulk bar > Assign role… > choose a role > Apply", "Role appears in the user's Roles badges after the list reloads"),
 ("Form guard", "Unsaved changes warning", "1. Edit a user; change Display Name without saving\n2. Click Users in the side nav\n3. Cancel the dialog, then try again and accept", "Browser confirm 'unsaved changes' appears; Cancel stays on the edit page; OK leaves without saving"),
 ("Idle warning", "Still-there dialog", "1. Stay fully idle (no mouse/keyboard) until SESSION_TIMEOUT_MINS minus 5 minutes elapse (testers: temporarily lower the timeout)\n2. Click 'Stay signed in'", "A 'Still there?' dialog appears before the session dies; Extend keeps the session alive and closes the dialog"),
 ("Template lifecycle", "Clone live template as draft", "1. System > Approval Templates\n2. Search UAT_WAVE_FLOW\n3. Click the clone icon on the ACTIVE row", "Toast 'Draft created…'; a new row UAT_WAVE_FLOW~D with a Draft badge appears; live row unchanged"),
 ("Template lifecycle", "Edit draft steps", "1. Open the draft (pencil icon)\n2. Rename step 1 and move it down\n3. Save, reopen", "Steps are editable only on drafts; rename + reorder persist after reopening"),
 ("Template lifecycle", "Activate draft", "1. Click Activate on the draft row\n2. Confirm", "Toast 'Draft activated — previous version archived'; draft becomes the ACTIVE UAT_WAVE_FLOW (version +1); old row archived as UAT_WAVE_FLOW~V1"),
 ("Audit log", "Before/after diff row", "1. System > Audit Log\n2. Click any row", "Row expands with a lazy-loaded panel: either a Field/Before/After diff table or the message that no snapshot was recorded"),
 ("Audit log", "Export CSV", "1. Audit Log > Export CSV", "A .csv file downloads containing the filtered audit entries"),
 ("Dashboard", "Customize layout persists", "1. Dashboard > Customize\n2. Hide one stat card, click Done\n3. Reload the app", "Hidden card stays hidden after reload (preference saved per user); Customize > re-enable restores it"),
 ("Notifications", "Unread badge count endpoint", "1. Note the bell badge number\n2. Compare with GET /dct/notifications/count", "Badge equals the endpoint's count; badge refreshes within ~60s of new notifications without a reload"),
 ("Audit info", "Created/Updated meta on edit pages", "1. Open any existing user in Edit\n2. Scroll to the bottom", "An audit line shows Created by/at and Updated by/at when available"),
]),
("Enhancement Round 2", "ENH", [
 ("Feature flags", "FEATURES render as toggles", "1. System Settings > FEATURES card\n2. Flip a switch (e.g. FEATURE_IDLE_WARNING) and Save\n3. Verify, then flip it back and Save", "FEATURE_* rows render as on/off switches (not text boxes); the flip persists and the feature honours it after reload"),
 ("Landing page", "Role edit landing dropdown", "1. Edit a role (e.g. MANAGER)\n2. Pick a Landing page from the dropdown and Save\n3. Restore Default afterwards", "Saving writes the LANDING_<ROLE> system setting; users with that role land on the chosen page at next login"),
 ("Audit log", "Date-range filter", "1. Audit Log: set From and To to today\n2. Set From to a future date", "Today's range shows entries (server-filtered count in the pager); a future From date shows an empty list"),
 ("Audit log", "Server-side CSV export", "1. Audit Log > Export CSV", "The CSV is built by the server over the FULL filtered history (up to 20k rows) — not capped at the first 1000 rows"),
 ("Sessions", "Active sessions page", "1. System > Active Sessions", "All sessions inside the inactivity window are listed with masked ids (…last6), login/last-activity times and IP; your own row is badged 'this device'"),
 ("Sessions", "Revoke signs the user out", "1. Login as a test user in another browser\n2. In Active Sessions click Revoke on that user and confirm", "All of that user's sessions close; their next click/navigation bounces them to the login page"),
 ("Bulk actions", "Bulk remove role", "1. Users: select a test user\n2. Bulk bar > Remove role… > pick the role > Apply", "The role disappears from the user's badges; other roles are untouched"),
 ("Bulk actions", "Bulk set organisation", "1. Select a test user\n2. Bulk bar > Set organisation… > pick an org > Apply", "The Organisation column updates for the selected user(s)"),
 ("Command palette", "Action verbs", "1. Ctrl+K\n2. Type 'create user' and pick the action", "An Actions group offers Create user / Create role / New announcement / New delegation; Create user opens a blank Add User form"),
 ("Command palette", "Recent pages", "1. Visit Users, then Roles, then Home\n2. Press Ctrl+K with an empty query", "A Recent group lists the pages just visited; picking one navigates back to it"),
 ("Template history", "Version history + step diff", "1. Approval Templates: search UAT_WAVE_FLOW\n2. Click the history (🕘) icon on the ACTIVE row", "Archived ~V versions no longer clutter the main list; the History modal shows live + archived versions with step-diff chips (added/removed/reordered vs live)"),
 ("Template history", "Restore archived version", "1. In the History modal click Restore on an archived version\n2. Confirm", "A new ~D draft is created with the archived steps; the live template is unchanged until the draft is activated"),
 ("Module rollout", "Palette + idle warning in module apps", "1. Open Petty Cash (or DT/HR/FL/CC) via the switcher\n2. Press Ctrl+K\n3. Stay idle to see the warning (testers: lower the timeout)", "The command palette opens with that app's navigation; the idle 'Still there?' warning and unsaved-changes guard are active in all module apps"),
]),
]

PC = [
("Access & Session", "ACC", [
 ("Shared session", "Entry from Admin", "1. Login in the Admin app\n2. Use the module switcher to open Petty Cash", "PC opens logged-in (no second login); your name in the top bar"),
 ("No session", "Direct open without login", "1. Logout (or new private window)\n2. Open the PC app URL directly", "Redirected to the Admin login page"),
 ("Role gating", "Approver/Admin menus", "1. Login as a regular user, then as a PC admin", "Approvals group only for approvers; All Petty Cash/Configuration only for admins"),
]),
("Dashboard", "DSH", [
 ("Stats", "Cards show live numbers", "1. Open PC Dashboard", "Cards (my floats, pending, totals) show sensible non-error values"),
 ("Charts", "Floats by org + ageing", "1. Wait for both charts", "Outstanding floats by organisation bar chart + TEMPORARY ageing chart render correctly"),
 ("Charts", "Re-render on return", "1. Navigate away and back", "Charts render again, not blank/duplicated"),
]),
("My Petty Cash", "MPC", [
 ("My list", "My requests only", "1. Open My Petty Cash", "Only YOUR petty cash requests, with number, type, amount, status badge"),
 ("Create", "PERMANENT float request", "1. New Petty Cash Request\n2. Type=PERMANENT, amount, justification\n3. Add GL coding line(s) summing to the amount\n4. Save draft", "Draft created with number PC-YYYY-NNNNN; appears in My Petty Cash as DRAFT"),
 ("Create", "TEMPORARY float request", "1. Same with Type=TEMPORARY + needed-by/return date", "Draft created; type shown correctly"),
 ("Validation", "Coding must equal amount", "1. Create a request where coding lines total <> requested amount\n2. Try to submit", "Submission blocked with a clear message"),
 ("Validation", "Required fields", "1. Try saving with empty amount / justification", "Field-level validation errors; nothing saved"),
 ("Submit", "Send for approval", "1. Open a DRAFT request\n2. Click Submit", "Status changes to submitted/pending; appears in the approver's queue; status history records the event"),
 ("Detail", "Request detail", "1. Open a request from the list", "Full detail: amounts, coding lines, status history timeline, documents"),
 ("Withdraw", "Withdraw a submitted request", "1. Open your submitted request\n2. Withdraw", "Status returns to draft/withdrawn; it leaves the approver queue"),
]),
("Reimbursements", "RMB", [
 ("Create", "New reimbursement", "1. Open Reimbursements > New\n2. Add expense lines (date, GL code, amount, description)\n3. Save + Submit", "Reimbursement created with its own number; goes to approval"),
 ("Detail", "Line items + history", "1. Open a reimbursement", "Lines, totals, status history all shown"),
 ("My list", "Status tracking", "1. Check list after submit/approval", "Status badge updates through the lifecycle"),
]),
("Clearing", "CLR", [
 ("Create", "Clear a TEMPORARY float", "1. Open Clearing > New for a disbursed TEMPORARY float\n2. Add receipts/expense lines + returned amount\n3. Submit", "Clearing created; float and clearing linked; goes to approval"),
 ("Validation", "Totals must reconcile", "1. Make receipts + returned cash <> float amount\n2. Try to submit", "Blocked with a clear reconciliation message"),
 ("Detail", "Clearing detail", "1. Open a clearing record", "Receipt lines, float reference, status history shown"),
]),
("Approvals", "APR", [
 ("Queue", "Pending list (approver)", "1. Login as approver\n2. Open Approvals", "All PC items awaiting YOUR step; badge count matches"),
 ("Approve", "Approve with comment", "1. Open an item > Approve (+ comment)", "Item leaves queue; requester notified; history shows approver + timestamp + comment"),
 ("Reject", "Reject with reason", "1. Reject another item with a reason", "Status rejected; requester notified with the reason"),
 ("Authorization", "Non-approver blocked", "1. Login as a regular user\n2. Try the Approvals route", "No Approvals menu; direct navigation shows no queue/denied"),
]),
("Administration", "ADM", [
 ("All Petty Cash", "Paged admin list", "1. Open All Petty Cash (admin)\n2. Check pager '1-50 of N'\n3. Search + filter by status", "All requests org-wide; server pagination, search and status filter work together"),
 ("All Reimbursements", "Admin list", "1. Open All Reimbursements", "Org-wide list with status filters"),
 ("All Clearings", "Admin list", "1. Open All Clearings", "Org-wide list; links open details"),
 ("Disburse", "Mark approved float disbursed", "1. Open an APPROVED request\n2. Disburse (date/reference)", "Status becomes DISBURSED; history updated; TEMPORARY floats start ageing"),
]),
("Configuration", "CFG", [
 ("GL codes", "Browse + search combinations", "1. Open GL Codes\n2. Search a segment value", "9-segment combinations listed with full code string; search filters"),
 ("Approval rules", "View/edit PC rules", "1. Open Approval Rules\n2. Review steps/amount thresholds", "Rules display matches the configured approval template"),
 ("Module settings", "PC settings incl. brand", "1. Open Module Settings\n2. Edit a setting (e.g. THEME_BRAND_COLOR)\n3. Hard-refresh", "Setting persists; brand color change reflects in PC only (restore afterwards)"),
]),
("Shell & Language", "SHL", [
 ("Switcher", "Back to Admin / other modules", "1. Use the top-bar module dropdown", "Other apps open without re-login"),
 ("Arabic", "RTL pass", "1. Switch to AR\n2. Walk My Petty Cash + Dashboard", "Shell + navigation in Arabic, layout mirrored, amounts keep Latin digits; lists still load"),
 ("Notifications", "PC notifications", "1. Trigger an approval event\n2. Check the bell", "Notification received and readable"),
]),
]

DT = [
("Access & Session", "ACC", [
 ("Shared session", "Entry from Admin", "1. Login in Admin\n2. Switch to Duty Travel via the module dropdown", "DT opens logged-in; correct nav for your role"),
 ("No session", "Direct open without login", "1. In a fresh private window open the DT URL", "Redirected to Admin login"),
 ("Role gating", "Finance/Admin menus", "1. Compare menus as requester vs finance vs admin", "Finance (disbursement/closure) and Admin (all requests/report) groups only for those roles"),
]),
("Dashboard", "DSH", [
 ("Stats", "Cards + recent requests", "1. Open DT Dashboard", "Cards show counts; recent requests list populated"),
 ("Charts", "Monthly spend + status funnel", "1. Wait for both charts", "12-month advances vs per-diem line + status funnel render with data"),
]),
("My Travel", "TRV", [
 ("My requests", "List + status", "1. Open My Requests", "Your travel requests with destination, dates, amount, status badge"),
 ("Create", "New travel request", "1. New Request: traveller, destination country/city, purpose, dates\n2. Check per-diem auto-calculation\n3. Save draft", "Per-diem computes from country group + grade and nights; draft saved with request number"),
 ("Create", "Advance request", "1. In the request enable/enter a cash advance", "Advance amount validated against policy and stored"),
 ("Validation", "Dates + required fields", "1. Try end date before start date; missing destination", "Clear validation errors; cannot submit"),
 ("Documents", "Attach required docs", "1. Attach the documents required for the destination/type\n2. Try submitting without them", "Missing mandatory docs block submission; with docs it submits"),
 ("Submit", "Send for approval", "1. Submit the draft", "Status pending; appears in approver queue; status history started"),
 ("Detail", "Request detail", "1. Open a request", "Segments, per-diem breakdown, advance, documents, history timeline all visible"),
 ("Settlement", "Create settlement", "1. After a disbursed trip open My Settlements > New\n2. Enter actuals/receipts\n3. Submit", "Settlement created and linked to the request; difference (refund/claim) computed"),
 ("My settlements", "Track settlement status", "1. Open My Settlements", "Settlements with statuses; detail opens"),
]),
("Approvals", "APR", [
 ("Queue", "Pending approvals", "1. As approver open Approvals", "DT items awaiting your step; badge matches"),
 ("Approve / Reject", "Action with comment", "1. Approve one; reject one with reason", "Statuses advance; requester notified; history updated"),
]),
("Finance", "FIN", [
 ("Disbursement queue", "Pay approved advances", "1. As finance open Disbursement Queue\n2. Mark an approved request disbursed", "Request leaves the queue; status DISBURSED; history updated"),
 ("Closure queue", "Close settled trips", "1. Open Closure Queue\n2. Close a fully settled request", "Request closed; removed from open lists"),
 ("All settlements", "Finance settlements view", "1. Open All Settlements", "Org-wide settlements with statuses and amounts"),
]),
("Administration", "ADM", [
 ("All requests", "Paged admin list", "1. Open All Requests\n2. Check pager; search; filter by status", "Org-wide server-paged list; filters work; 'mine' lists elsewhere unaffected"),
 ("Travel report", "Aggregated report", "1. Open Travel Report\n2. Apply period/org filters", "Aggregates (trips, spend, advances) consistent with the lists"),
]),
("Configuration", "CFG", [
 ("Per-diem rates", "Rates by group/grade", "1. Open Per Diem Rates\n2. Edit one rate\n3. Create a new request for that group/grade", "New rate persists and is used in the calculation (restore afterwards)"),
 ("Country groups", "Country assignment", "1. Open Country Groups\n2. Move a country to another group", "Mapping persists; per-diem for that country follows the new group"),
 ("Doc requirements", "Required docs matrix", "1. Open Doc Requirements\n2. Add a required doc for a travel type", "New requirement enforced on the next request of that type (remove afterwards)"),
 ("Approval rules", "DT rules", "1. Open Approval Rules", "Steps/thresholds match the approval template"),
 ("Module settings", "DT settings", "1. Edit a DT module setting and refresh", "Value persists; behaviour/brand follows"),
]),
("Shell & Language", "SHL", [
 ("Switcher", "Cross-module hop", "1. Use the module dropdown to another app and back", "No re-login; DT state intact"),
 ("Arabic", "RTL pass", "1. Switch to AR and walk My Requests + Dashboard", "Mirrored layout, Arabic shell labels, Latin digits, lists load"),
 ("Notifications", "DT notifications", "1. Trigger approval/disbursement events", "Bell shows them; readable detail"),
]),
]

HR = [
("Access & Session", "ACC", [
 ("Shared session", "Entry from Admin", "1. Login in Admin\n2. Switch to HR via the module dropdown", "HR opens logged-in; viewer/manager/admin nav per role"),
 ("No session", "Direct open without login", "1. Fresh private window > HR URL", "Redirected to Admin login"),
]),
("Dashboard", "DSH", [
 ("Stats", "Headcount cards", "1. Open HR Dashboard", "Cards (headcount, vacancies, expiring docs) show live values"),
 ("Charts", "Headcount + expiry horizon", "1. Wait for both charts", "Approved vs filled headcount chart + document expiry horizon chart render"),
]),
("People", "EMP", [
 ("Employees list", "Paged list", "1. Open Employees\n2. Page next/previous", "Employee cards/rows with photo or initials; paging works without losing filters"),
 ("Employees list", "Search + filters", "1. Search by name\n2. Filter by organisation and grade", "Server-side filtering; combined filters work; clear restores"),
 ("Employee detail", "Full record", "1. Open an employee", "Personal, job (position/grade/org), contacts, documents tabs all populated"),
 ("Create", "New employee", "1. New Employee: number, names EN/AR, hire date, position, org, grade\n2. Save", "Employee created and findable; appears in headcount"),
 ("Edit", "Update employee", "1. Edit job title / phone / org\n2. Save and reopen", "Changes persist"),
 ("Photo", "Upload employee photo", "1. In employee detail click photo/upload\n2. Choose an image", "Picker opens; photo saved and shown in list + detail (small file < 20KB if large files fail)"),
 ("Documents", "Add employee document", "1. Add a document (type, number, expiry date) to an employee\n2. Upload the file", "Document listed with expiry; appears in compliance views"),
 ("Org hierarchy", "Org chart", "1. Open Org Hierarchy", "Tree/chart of org units with employees; expand/collapse works"),
]),
("Structure", "STR", [
 ("Positions", "CRUD", "1. Open Positions\n2. Create a position (title, org, grade, headcount)\n3. Edit it\n4. Try assigning an employee to it", "Position created/edited; assignable to employees; filled count updates"),
 ("Jobs", "CRUD", "1. Open Jobs\n2. Create/edit a job", "Job persists; selectable on positions/employees"),
 ("Grades", "List + edit", "1. Open Grades\n2. Edit a grade label", "Grades listed; edit persists; grade dropdowns reflect it"),
 ("Locations", "CRUD", "1. Open Locations\n2. Add/edit a location", "Location persists; selectable on employees/positions"),
]),
("Compliance", "CMP", [
 ("Expiring documents", "Expiry list + badge", "1. Open Documents (compliance)\n2. Compare the side-nav badge number with rows", "Documents expiring within the window listed with days-left; badge matches"),
 ("Document detail", "Navigate to employee", "1. Click an expiring document", "Opens the owning employee's documents tab"),
 ("Renewal", "Update expiry", "1. Update a document's expiry to a future date", "Item leaves the expiring list; badge decreases"),
]),
("Administration", "ADM", [
 ("Lookups", "HR lookup values", "1. Open Lookups\n2. Add/edit a value in an HR category", "Value persists and appears in the matching dropdowns"),
 ("Module settings", "HR settings", "1. Open Module Settings\n2. Edit a value and refresh", "Persists; no errors"),
]),
("Shell & Language", "SHL", [
 ("Switcher", "Cross-module hop", "1. Module dropdown > Admin and back", "No re-login"),
 ("Arabic", "RTL pass", "1. Switch to AR\n2. Walk Employees + Dashboard", "Mirrored layout; Arabic shell; Arabic names display correctly; Latin digits"),
 ("Notifications", "HR notifications", "1. Check the bell after document-expiry alerts run", "Expiry notifications visible and readable"),
]),
]

FL = [
("Access & Session", "ACC", [
 ("Shared session", "Entry from Admin", "1. Login in Admin\n2. Switch to Freelancers via the module dropdown", "FL opens logged-in; viewer/manager/admin nav per role (FL_USER / FL_MANAGER / FL_ADMIN)"),
 ("No session", "Direct open without login", "1. Fresh private window > FL URL", "Redirected to Admin login"),
 ("Role gating", "Compliance/Approvals/Admin menus", "1. Compare menus as FL_USER (SHAIKHA.GALAMERI) vs FL_MANAGER (NASER) vs FL_ADMIN (AYESHA)", "Compliance + Approvals groups appear for managers+; Module Settings only for FL_ADMIN/SYS_ADMIN"),
]),
("Dashboard", "DSH", [
 ("Stats", "Cards show live numbers", "1. Open FL Dashboard", "Active freelancers, active contracts (+AED total), vouchers awaiting approval, documents expiring — all sensible values"),
 ("Charts", "Spend + expiry horizon", "1. Wait for both charts", "Committed-vs-Paid by organisation bar chart + document expiry horizon chart render with data"),
]),
("Registrations", "REG", [
 ("List", "Paged list + search", "1. Open Registrations\n2. Search by name\n3. Filter by status", "Server-paged list; filters work; FL-REG numbers shown"),
 ("Create", "New registration draft", "1. + New Registration: names, date of birth, nationality, email, mobile, specialization\n2. Save", "Draft saved with FL-REG-NNNNNN number; appears in the list as DRAFT"),
 ("Photo", "Photo required to submit", "1. Open a DRAFT registration without a photo\n2. Try Submit\n3. Upload a photo and submit again", "Submit blocked until a photo is uploaded; with photo it submits"),
 ("Submit", "Send for approval", "1. Submit a complete draft", "Status SUBMITTED; appears in the FL manager's approvals queue"),
 ("Approve", "Approval creates the freelancer", "1. As FL_MANAGER approve the registration (then FL_ADMIN if a second step shows)", "On final approval a freelancer profile is created automatically with a FL-VND vendor number"),
]),
("Freelancers", "VND", [
 ("List", "Directory", "1. Open Freelancers", "Active freelancers with vendor number, name, specialization, status"),
 ("Detail", "Tabs", "1. Open a freelancer", "Profile / Bank Accounts / Contracts / Documents tabs all load; audit info shows created/updated in local time"),
 ("Bank", "Add bank account", "1. In Bank Accounts add IBAN/bank details\n2. Save", "Account listed; primary flag honoured"),
 ("Status", "Change status", "1. Use Change Status (e.g. ACTIVE > INACTIVE and back)", "Status updates with a reason; reflected in the list"),
 ("Documents", "Add document with expiry", "1. In Documents add a document with a near-future expiry\n2. Upload the file", "Document saved; appears in Compliance > Documents with days-to-expiry"),
]),
("Contracts", "CON", [
 ("List", "Contracts with billing bar", "1. Open Contracts", "Contracts with totals, status and % billed progress bar"),
 ("Create", "New contract draft", "1. + New Contract: freelancer, title, dates, total amount, billing method, GL or Project coding\n2. Save", "Draft saved with FL-CON number; GL/PROJECT coding toggle works"),
 ("Submit small", "< 50k = 1 approval step", "1. Submit a contract under AED 50,000\n2. Approve as FL_MANAGER", "Contract becomes ACTIVE after ONE approval; payment schedule generated automatically"),
 ("Submit large", ">= 50k = 2 approval steps", "1. Submit a contract of AED 50,000+ (seeded: FL-CON-000002, 84k)\n2. Approve as FL_MANAGER", "After step 1 it is still pending at the FL_ADMIN step; only after step 2 does it activate"),
 ("Detail", "Schedule + tabs", "1. Open an ACTIVE contract", "Payment Schedule / Amendments / Renewals / Details tabs; schedule rows match the billing method"),
 ("Amendment", "Amend amount/end date", "1. + Amendment on an ACTIVE contract\n2. Submit and approve it", "On approval the contract values update, version +1, schedule rebuilt"),
 ("Renewal", "Renew a contract", "1. Renew from the contract detail\n2. Submit and approve", "A new contract version/row is created per the renewal terms"),
]),
("Payments", "PAY", [
 ("Schedule", "Due worklist", "1. Open Payment Schedule", "Due/pending schedule rows across contracts with Generate Voucher actions"),
 ("Voucher", "Generate from schedule", "1. Click Generate Voucher on a due row", "Voucher FL-VCH created in DRAFT linked to the schedule row"),
 ("Voucher", "Invoice + submit", "1. Open the voucher detail\n2. Enter invoice reference\n3. Submit and approve", "Voucher APPROVED; schedule row marked voucher-generated"),
 ("Mark paid", "FL_ADMIN settles", "1. As FL_ADMIN use Mark Paid on an approved voucher", "Payment status PAID; contract paid-total and % billed update"),
 ("Deliverables", "Accept / reject", "1. Open Deliverables\n2. Accept one and reject another with a reason", "States update; reason stored"),
]),
("Compliance", "CMP", [
 ("Documents", "Expiry filter + badge", "1. Open Compliance > Documents\n2. Use the expiry filter (Expired / Expiring Soon / Valid)", "Filtered list with days-to-expiry; side-nav badge matches the expiring count"),
 ("Navigate", "Document to freelancer", "1. Click a document row", "Opens the owning freelancer's Documents tab"),
]),
("Approvals", "APR", [
 ("Queue", "Pending list", "1. Login as FL_MANAGER (NASER)\n2. Open Pending Approvals", "FL registrations/contracts/vouchers awaiting YOUR step; badge count matches"),
 ("Comment required", "Approve needs a comment", "1. Click Approve and try to confirm with an empty comment", "Blocked with 'a comment is required'; with a comment it succeeds"),
 ("Reject / Return", "Negative paths", "1. Reject one item; Return another (with comments)", "Statuses update; requester notified; returned item editable again"),
]),
("Administration", "ADM", [
 ("Module settings", "FL settings", "1. As FL_ADMIN open Module Settings\n2. Edit a value (e.g. DOC_EXPIRY_ALERT_DAYS)\n3. Save + refresh", "Persists; non-admins do not see the page"),
]),
("Shell & Language", "SHL", [
 ("Switcher", "Cross-module hop", "1. Module dropdown > Admin / CC and back", "No re-login; FL shows live (no 'soon' badge)"),
 ("Arabic", "RTL pass", "1. Switch to AR\n2. Walk Dashboard + Registrations + a contract detail", "Mirrored layout, Arabic labels, Latin digits; charts re-render"),
]),
("Enhancements", "ENH", [
 ("Compliance deep link", "Doc row opens Documents tab", "1. Open Compliance > Documents\n2. Click a document row", "Freelancer detail opens directly on the Documents tab (not Profile) — FL-CMP-02 fix"),
 ("Bulk kill switch", "Bulk button respects setting", "1. With ALLOW_BULK_VOUCHER_GENERATION = N open Payment Schedule\n2. Call POST /fl/schedule/bulk-generate directly", "No 'Generate All Due' button on the page; direct API call refused with 403"),
 ("Bulk generation", "Generate all due vouchers", "1. Set ALLOW_BULK_VOUCHER_GENERATION = Y\n2. Reload Payment Schedule\n3. Generate All Due > Generate\n4. Restore the setting to N", "Button appears; one DRAFT voucher per PENDING row without an open voucher; rows flip to VOUCHER_GENERATED; toast shows the created count"),
 ("Unsaved changes guard", "Dirty form warns on navigation", "1. Open a contract/registration edit form\n2. Change a field without saving\n3. Click another nav item", "Confirm dialog about unsaved changes; Cancel stays, OK leaves; after Save no warning"),
 ("Command palette", "Ctrl+K navigation", "1. Press Ctrl+K anywhere in FL\n2. Type a page name and press Enter", "Palette opens with FL navigation entries; Enter navigates to the page"),
]),
]

CC = [
("Access & Session", "ACC", [
 ("Shared session", "Entry from Admin", "1. Login in Admin\n2. Switch to Credit Cards via the module dropdown", "CC opens logged-in; holder vs approver vs CC_ADMIN nav per role"),
 ("No session", "Direct open without login", "1. Fresh private window > CC URL", "Redirected to Admin login"),
 ("Role gating", "Admin menus", "1. Compare menus as a holder (SHAIKHA.ALSUWAIDI) vs CC_ADMIN (NASER)", "All Cards / Proxies / Module Settings only for CC_ADMIN/SYS_ADMIN; Approvals for MANAGER+"),
]),
("Dashboard", "DSH", [
 ("Stats", "Cards show live numbers", "1. Open CC Dashboard", "Active cards, total credit limit, pending requests, replenishments due — sensible values"),
 ("Charts", "Limits + compliance", "1. Wait for both charts", "Credit limits by organisation bar + 6-month replenishment compliance (submitted vs due) render"),
]),
("My Card", "CRD", [
 ("Card visual", "My card renders", "1. Login as a card holder\n2. Open My Card", "Card visual with number, holder, limit, expiry; status badge; greyed if not ACTIVE"),
 ("Limit history", "Timeline", "1. Check the Limit History table", "INITIAL row from registration; INCREASE/DECREASE rows after approved limit changes"),
 ("Due banner", "Replenishment due", "1. As a holder whose current-month statement is missing open My Card", "Amber 'this month's replenishment has not been submitted' banner with a shortcut button"),
 ("No card", "Holder without card", "1. Login as a user with no card\n2. Open My Card", "Friendly empty state offering a New Card request (no errors)"),
]),
("Card Requests", "REQ", [
 ("Wizard", "Type tiles", "1. Open Card Requests > + New Request", "5 tiles: New Card / Increase / Decrease / Replacement / Close; 3-step wizard with progress"),
 ("New card", "Full happy path", "1. Pick New Card, enter limit + reason\n2. Save Draft & Continue\n3. Upload every mandatory document\n4. Submit", "Submit only enabled once all mandatory docs are uploaded; request goes to approval"),
 ("Doc checklist", "Mandatory docs enforced", "1. In step 3 leave a mandatory document missing\n2. Try Submit", "Submit disabled / blocked with a clear message naming the missing document"),
 ("Increase limit", "Validation", "1. Create INCREASE_LIMIT with a limit LOWER than the current one\n2. Submit", "Blocked: requested limit must exceed the current limit"),
 ("Withdraw", "Pull back a submitted request", "1. Open your SUBMITTED request\n2. Withdraw", "Status WITHDRAWN; leaves the approver queue; card free for new requests"),
 ("Detail", "Request detail", "1. Open a request", "Type, card, limits, reason, documents and audit info all shown"),
]),
("Replenishments", "RPL", [
 ("Create", "New monthly statement", "1. Open Replenishments > + New Replenishment\n2. Pick your card + last month\n3. Create Draft", "Draft CCM-YYYY-MM-NNNNN created; one statement per card per month enforced"),
 ("Lines", "Merchant lines editor", "1. In the detail add 3 lines: date, merchant, amount, GL or Project coding, receipt flag\n2. Save Lines", "Lines saved; running total recalculates; receipt counter updates"),
 ("Validation", "Line rules", "1. Try saving a line without merchant/amount/date or GL code", "Clear per-line validation; nothing saved"),
 ("Submit", "Send for approval", "1. Submit the statement", "Status SUBMITTED; appears in the approvals queue; only the holder or an active proxy may submit"),
 ("Proxy", "Proxy submits for the holder", "1. As CC_ADMIN create a proxy for a card\n2. Login as the proxy and create/submit a statement for that card", "Proxy can submit; non-proxies are blocked with a clear message"),
]),
("Approvals", "APR", [
 ("Queue", "Pending list", "1. Login as an approver (MANAGER/CC_ADMIN)\n2. Open Pending Approvals", "CC requests + replenishments awaiting your step; badge matches"),
 ("Comment required", "Approve needs a comment", "1. Approve with an empty comment", "Blocked; succeeds with a comment"),
 ("Lifecycle", "Approved request applies", "1. Fully approve an INCREASE_LIMIT request", "Card limit updates; INCREASE row added to limit history; holder notified"),
]),
("Administration", "ADM", [
 ("All cards", "Registry + search", "1. As CC_ADMIN open All Cards\n2. Search and filter by status", "All cards org-wide; pending operation chips for in-flight requests"),
 ("Register card", "From approved NEW_CARD", "1. + Register Card\n2. Pick an APPROVED NEW_CARD request\n3. Fill holder, mother name, nationality, limit, expiry, email, org\n4. Register", "Card created ACTIVE with CC-YYYY-NNNNN number + INITIAL limit history; holder notified"),
 ("Proxies", "Manage proxies", "1. Open Proxies\n2. Create one (card + proxy user + window)\n3. Deactivate it", "Proxy listed; deactivation removes submit rights"),
 ("Module settings", "CC settings", "1. Open Module Settings\n2. Edit a value (e.g. REPLENISHMENT_DUE_DAY)\n3. Save + refresh", "Persists; non-admins do not see the page"),
]),
("Shell & Language", "SHL", [
 ("Switcher", "Cross-module hop", "1. Module dropdown > Admin / FL and back", "No re-login; CC shows live (no 'soon' badge)"),
 ("Arabic", "RTL pass", "1. Switch to AR\n2. Walk Dashboard + My Card + a request detail", "Mirrored layout, Arabic labels, Latin digits; card visual unaffected"),
]),
]

APPS = [
    ("Admin", "Admin (App 200) — Identity, Users, Roles, Platform Administration", ADMIN),
    ("PC",    "Petty Cash (App 201) — Floats, Reimbursements, Clearing",            PC),
    ("DT",    "Duty Travel (App 204) — Travel Requests, Per-Diem, Settlements",     DT),
    ("HR",    "HR (App 205) — Employees, Structure, Compliance",                    HR),
    ("FL",    "Freelancers (App 203) — Registrations, Contracts, Payments",         FL),
    ("CC",    "Credit Cards (App 202) — Cards, Requests, Replenishments",           CC),
]

KNOWN_ISSUES = {
    "HR": [],
    "Admin": [],
    "PC": [],
    "DT": [],
    "FL": [],
    "CC": [],
}

SEED_NOTES = {
    "Admin": [
        "6 test accounts exist matching the quick-login buttons: ADMIN (SYS_ADMIN), HASHEM.ALKABBI "
        "(Director/approver), NASER.ALKHAJA, AYESHA.AMERI, SHAIKHA.GALAMERI, SHAIKHA.ALSUWAIDI (managers).",
        "Module roles: AYESHA = Petty Cash admin/approver, HASHEM = DT approver, NASER = DT finance, "
        "SHAIKHA.GALAMERI = HR manager, SHAIKHA.ALSUWAIDI = HR viewer.",
    ],
    "PC": [
        "PC-2026-00008 (SHAIKHA.ALSUWAIDI, TEMPORARY 1,800) - awaiting approval: use for approve/reject tests (approver: AYESHA.AMERI).",
        "PC-2026-00005 (NASER, TEMPORARY 2,500) - fully approved, awaiting disbursement: use for the Disburse test.",
        "PC-2026-00006 (AYESHA, TEMPORARY 3,500) - ACTIVE/disbursed: use to create a Clearing.",
        "PC-2026-00007 (SHAIKHA.GALAMERI, PERMANENT 5,000) - ACTIVE: use to create Reimbursements.",
        "RMB-2026-00100 (450 on PC-00007) - awaiting approval.",
    ],
    "DT": [
        "DT-REQ-2026-00001 (NASER, Paris) - DRAFT: use for edit/submit tests.",
        "DT-REQ-2026-00002 (AYESHA, London) - SUBMITTED: in HASHEM's approval queue.",
        "DT-REQ-2026-00003 (SHAIKHA.GALAMERI, Riyadh, trip already past) - ADVANCE_PAID: use to create a Settlement.",
        "Config seeded: per-diem rates (18 countries + tiers + regions), 50 country-group mappings, 2 approval workflows.",
    ],
    "HR": [
        "5 active employees incl. E1101 Mohammed Al Mansoori, E1102 Fatima Al Hosani, E1103 Omar Khalil (seeded).",
        "Documents seeded with expiries ~20 / ~45 days (compliance list + badge data) and one ~13 months out.",
        "HR roles: SHAIKHA.GALAMERI = HR_MANAGER, SHAIKHA.ALSUWAIDI = HR_VIEWER.",
    ],
    "FL": [
        "FL roles: AYESHA.AMERI = FL_ADMIN, NASER.ALKHAJA = FL_MANAGER, SHAIKHA.GALAMERI = FL_USER.",
        "Layla Hassan (FL-VND-000001) - ACTIVE freelancer with bank account, documents and contract FL-CON-000001 "
        "(24k, ACTIVE, 1 voucher PAID): use for detail/schedule/voucher tests.",
        "FL-CON-000002 (84k) - pending at the FL_ADMIN step: use for the >=50k two-step approval test.",
        "FL-REG-000004 (Omar Saleh) - DRAFT; FL-REG-000005 (Mariam Khalil) - SUBMITTED, in the manager queue.",
        "One amendment, one deliverable, one renewal and near-expiry documents exist from the lifecycle seed.",
    ],
    "CC": [
        "CC roles: NASER.ALKHAJA = CC_ADMIN; card holders: SHAIKHA.ALSUWAIDI (CC-2026-00002, 20k) and "
        "HASHEM.ALKABBI (CC-2026-00003, 12k). CC-2026-00001 is a CLOSED test card.",
        "CCM-2026-05-00001 (May statement, 3 merchant lines, 1,970.50) - SUBMITTED, awaiting approval.",
        "One INCREASE_LIMIT request exists in WITHDRAWN state (negative-path evidence).",
        "Doc requirements seeded: 4 mandatory documents for NEW_CARD; check the wizard checklist against them.",
    ],
}


# =============================================================================
# Sample data per test case — keyed (app, area prefix, scenario). Falls back to
# AREA_SAMPLES[(app, prefix)] then ''. References seeded records and usernames
# only — NEVER passwords (testers use the quick-login buttons).
# =============================================================================
AREA_SAMPLES = {
    ("Admin", "LOG"): "Quick-login buttons on the login page (ADMIN, HASHEM.ALKABBI, NASER.ALKHAJA…)",
    ("Admin", "DLG"): "Seeded: delegation NASER.ALKHAJA → AYESHA.AMERI (ACTIVE, 14 days); 2 announcements (INFO/ALL + WARNING/CREDIT_CARDS)",
    ("PC",  "APR"): "PC-2026-00008 (SHAIKHA.ALSUWAIDI, 1,800) awaiting approval — approver AYESHA.AMERI",
    ("DT",  "APR"): "DT-REQ-2026-00002 (AYESHA, London) — in HASHEM.ALKABBI's queue",
    ("HR",  "EMP"): "Seeded employees: E1101 Mohammed Al Mansoori, E1102 Fatima Al Hosani, E1103 Omar Khalil",
    ("FL",  "ACC"): "FL_USER = SHAIKHA.GALAMERI · FL_MANAGER = NASER.ALKHAJA · FL_ADMIN = AYESHA.AMERI",
    ("FL",  "APR"): "Login NASER.ALKHAJA; pending: FL-REG-000005 (registration) + FL-CON-000002 (84k contract, step 2)",
    ("CC",  "ACC"): "Holders: SHAIKHA.ALSUWAIDI (CC-2026-00002), HASHEM.ALKABBI (CC-2026-00003) · CC_ADMIN = NASER.ALKHAJA",
    ("CC",  "APR"): "CCM-2026-05-00001 (May statement, 1,970.50) — SUBMITTED, in the approver queue",
}

SAMPLES = {
    # ── Admin ────────────────────────────────────────────────────────────────
    ("Admin", "LOG", "Valid login"): "Quick-login: System Admin (ADMIN)",
    ("Admin", "LOG", "Invalid password"): "Username ADMIN + any wrong password",
    ("Admin", "LOG", "Inactive account"): "Create + deactivate UAT.TESTER1 first (User Management)",
    ("Admin", "WSP", "Update phone persists"): "Phone: +97150 000 1234 (restore the original after)",
    ("Admin", "WSP", "Upload photo"): "Any JPG/PNG photo file",
    ("Admin", "USR", "Happy path"): "Username UAT.TESTER1 · role PLATFORM_USER · org Finance Division",
    ("Admin", "USR", "Duplicate username"): "Existing username: NASER.ALKHAJA",
    ("Admin", "USR", "Change roles"): "Test user UAT.TESTER1 — add/remove AUDITOR",
    ("Admin", "ROL", "New role"): "Code UAT_ROLE1, name 'UAT Test Role' (delete/deactivate after)",
    ("Admin", "SYS", "Add lookup value"): "Category CC_REPLACEMENT_REASON — add code UAT_TEST",
    ("Admin", "SYS", "Brand color from settings"): "THEME_BRAND_COLOR = #7B1FA2 (restore #C74634 after)",
    ("Admin", "DLG", "Profile shows real data"): "Login AYESHA.AMERI — shows 'Acting for Naser Al Khaja' (incoming, seeded)",
    ("Admin", "DLG", "Delegate my authority"): "Delegate to HASHEM.ALKABBI, end date +7 days, reason 'UAT'",
    ("Admin", "DLG", "Delegate sees the queue"): "Login AYESHA.AMERI — NASER's pending items carry 'acting for' badge",
    ("Admin", "DLG", "Stop the cover"): "Cancel the delegation created in the previous case",
    ("Admin", "DLG", "System > Delegations"): "Seeded NASER→AYESHA row + your test rows",
    ("Admin", "DLG", "Create an announcement"): "Title 'UAT banner test' / INFO / Everyone (deactivate after)",
    ("Admin", "DLG", "Module-targeted banner"): "WARNING targeted at module CREDIT_CARDS (seeded example exists)",
    ("Admin", "DLG", "Dismiss persists for the session"): "Seeded banner: 'Freelancers & Credit Cards modules are live'",
    ("Admin", "DLG", "Pull an announcement"): "Deactivate your 'UAT banner test' announcement",
    ("Admin", "SHL", "FL + CC are live"): "Switcher: Freelancers APP 203 + Credit Cards APP 202 (no 'soon')",
    # ── PC ───────────────────────────────────────────────────────────────────
    ("PC", "MPC", "PERMANENT float request"): "Amount 2,000 · any active GL combination",
    ("PC", "MPC", "TEMPORARY float request"): "Amount 1,500 · needed-by +7 days",
    ("PC", "RMB", "New reimbursement"): "On PC-2026-00007 (SHAIKHA.GALAMERI, PERMANENT 5,000); RMB-2026-00100 (450) already awaiting approval",
    ("PC", "CLR", "Clear a TEMPORARY float"): "PC-2026-00006 (AYESHA, 3,500, ACTIVE/disbursed)",
    ("PC", "APR", "Pending list (approver)"): "Login AYESHA.AMERI — PC-2026-00008 (1,800) in the queue",
    ("PC", "ADM", "Mark approved float disbursed"): "PC-2026-00005 (NASER, 2,500) — approved, awaiting disbursement",
    # ── DT ───────────────────────────────────────────────────────────────────
    ("DT", "TRV", "List + status"): "Login NASER.ALKHAJA — DT-REQ-2026-00001 (Paris, DRAFT)",
    ("DT", "TRV", "New travel request"): "Destination London, 3 nights, next month",
    ("DT", "TRV", "Create settlement"): "DT-REQ-2026-00003 (SHAIKHA.GALAMERI, Riyadh, ADVANCE_PAID, trip past)",
    ("DT", "APR", "Pending approvals"): "Login HASHEM.ALKABBI — DT-REQ-2026-00002 (AYESHA, London) in the queue",
    ("DT", "FIN", "Pay approved advances"): "Login NASER.ALKHAJA (DT_FINANCE)",
    # ── HR ───────────────────────────────────────────────────────────────────
    ("HR", "EMP", "Full record"): "E1101 Mohammed Al Mansoori",
    ("HR", "EMP", "New employee"): "Number E9901 · hire date today · any position/org/grade",
    ("HR", "EMP", "Upload employee photo"): "E1102 Fatima Al Hosani · small JPG/PNG (< 20KB if large files fail)",
    ("HR", "CMP", "Expiry list + badge"): "Seeded documents expiring in ~20 and ~45 days",
    # ── FL ───────────────────────────────────────────────────────────────────
    ("FL", "REG", "Paged list + search"): "FL-REG-000004 (Omar Saleh, DRAFT) · FL-REG-000005 (Mariam Khalil, SUBMITTED)",
    ("FL", "REG", "New registration draft"): "Name 'Test Freelancer UAT' · nationality from dropdown · test.fl.uat@example.com",
    ("FL", "REG", "Photo required to submit"): "FL-REG-000004 (no photo yet) · any JPG/PNG",
    ("FL", "REG", "Send for approval"): "Your draft from the previous case (photo uploaded)",
    ("FL", "REG", "Approval creates the freelancer"): "FL-REG-000005 — in NASER.ALKHAJA's queue",
    ("FL", "VND", "Directory"): "Layla Hassan — FL-VND-000001 (ACTIVE)",
    ("FL", "VND", "Tabs"): "FL-VND-000001 (has bank account, contract, documents)",
    ("FL", "VND", "Add bank account"): "IBAN AE07 0331 2345 6789 0123 456 · any seeded bank",
    ("FL", "VND", "Add document with expiry"): "Expiry +20 days (so it appears in Compliance)",
    ("FL", "CON", "Contracts with billing bar"): "FL-CON-000001 (24k ACTIVE, ~17% billed) · FL-CON-000002 (84k, pending)",
    ("FL", "CON", "New contract draft"): "FL-VND-000001 · 12 months · GL coding from dropdown",
    ("FL", "CON", "< 50k = 1 approval step"): "Amount 24,000 — approve as NASER.ALKHAJA only",
    ("FL", "CON", ">= 50k = 2 approval steps"): "FL-CON-000002 (84k) — already at step 2; approve as AYESHA.AMERI (FL_ADMIN)",
    ("FL", "CON", "Schedule + tabs"): "FL-CON-000001 — 6 monthly rows, June PAID",
    ("FL", "CON", "Amend amount/end date"): "FL-CON-000001 — one amendment already exists as reference",
    ("FL", "PAY", "Due worklist"): "FL-CON-000001 schedule — July onwards PENDING",
    ("FL", "PAY", "Generate from schedule"): "July 2026 row of FL-CON-000001",
    ("FL", "PAY", "Invoice + submit"): "Invoice ref INV-UAT-001",
    ("FL", "PAY", "FL_ADMIN settles"): "Login AYESHA.AMERI — FL-VCH-000001 is the PAID reference",
    ("FL", "CMP", "Expiry filter + badge"): "Seeded near-expiry document on FL-VND-000001",
    ("FL", "ADM", "FL settings"): "DOC_EXPIRY_ALERT_DAYS (restore the value after)",
    # ── CC ───────────────────────────────────────────────────────────────────
    ("CC", "CRD", "My card renders"): "Login SHAIKHA.ALSUWAIDI — CC-2026-00002 (20,000 AED)",
    ("CC", "CRD", "Timeline"): "CC-2026-00002 — INITIAL 0→20,000 row from registration",
    ("CC", "CRD", "Replenishment due"): "SHAIKHA.ALSUWAIDI — May submitted, current month missing → banner shows",
    ("CC", "CRD", "Holder without card"): "Login SHAIKHA.GALAMERI (no card)",
    ("CC", "REQ", "Full happy path"): "Limit 10,000 · 4 mandatory NEW_CARD docs (any small PNG/PDF)",
    ("CC", "REQ", "Mandatory docs enforced"): "Leave 'Salary Certificate' (or any mandatory doc) missing",
    ("CC", "REQ", "Validation"): "CC-2026-00002 (current 20,000) — request 15,000 to test the block",
    ("CC", "REQ", "Pull back a submitted request"): "Create INCREASE_LIMIT 25,000 on CC-2026-00003, submit, then withdraw (a seeded WITHDRAWN example exists)",
    ("CC", "RPL", "New monthly statement"): "CC-2026-00002 + May 2026 → duplicate blocked; use the current month",
    ("CC", "RPL", "Merchant lines editor"): "Lines: Etisalat 420.50 / Amazon AE 1,310.00 / ADNOC 240.00 (as in CCM-2026-05-00001)",
    ("CC", "RPL", "Send for approval"): "Your current-month draft from the previous cases",
    ("CC", "RPL", "Proxy submits for the holder"): "Proxy HASHEM.ALKABBI on CC-2026-00002 (create in Proxies first)",
    ("CC", "APR", "Approved request applies"): "INCREASE_LIMIT on CC-2026-00003: 12,000 → 14,000",
    ("CC", "ADM", "Registry + search"): "CC-2026-00001 (CLOSED) · 00002 (SHAIKHA.ALSUWAIDI) · 00003 (HASHEM)",
    ("CC", "ADM", "From approved NEW_CARD"): "Approve a fresh NEW_CARD request first (wizard case), then register it here",
    ("CC", "ADM", "Manage proxies"): "Card CC-2026-00002 · proxy SHAIKHA.GALAMERI · window today → +30 days",
    ("CC", "ADM", "CC settings"): "REPLENISHMENT_DUE_DAY (restore the value after)",
}


def style_header(ws, brand):
    fill = PatternFill("solid", fgColor=brand)
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = fill
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def build_area_sheet(wb, area, prefix, app_key, cases, brand):
    ws = wb.create_sheet(area[:31])
    style_header(ws, brand)
    r = 2
    for i, (func, scen, steps, exp) in enumerate(cases, start=1):
        tid = "%s-%s-%02d" % (app_key.upper(), prefix, i)
        sample = SAMPLES.get((app_key, prefix, scen),
                             AREA_SAMPLES.get((app_key, prefix), ""))
        vals = [tid, func, scen, steps, sample, exp, "", "", "", ""]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True,
                                    horizontal="center" if col in (1, 7, 8, 9) else "left")
            c.border = BORDER
            if col == 1:
                c.font = Font(bold=True, color="555555")
            if col == 5 and v:
                c.font = Font(color="1A5276", italic=True)
        r += 1
    last = r - 1

    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Choose Pass, Fail, Blocked or N/A"
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add("%s2:%s%d" % (STATUS_COL, STATUS_COL, last))

    rng = "%s2:%s%d" % (STATUS_COL, STATUS_COL, last)
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'],
        fill=PatternFill("solid", fgColor="C8E6C9"), font=Font(color="1B5E20", bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'],
        fill=PatternFill("solid", fgColor="FFCDD2"), font=Font(color="B71C1C", bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'],
        fill=PatternFill("solid", fgColor="FFE0B2"), font=Font(color="E65100", bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"N/A"'],
        fill=PatternFill("solid", fgColor="ECEFF1"), font=Font(color="546E7A")))

    ws.auto_filter.ref = "A1:J%d" % last
    return area, last - 1


def build_instructions(wb, app_key, app_title, areas_counts, brand):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFG", (30, 16, 12, 12, 12, 12)):
        ws.column_dimensions[col].width = w

    ws["B2"] = "i-Finance UAT Script"
    ws["B2"].font = Font(size=18, bold=True, color=brand)
    ws["B3"] = app_title
    ws["B3"].font = Font(size=12, bold=True, color="444444")

    meta = [("Environment URL", "(fill in: e.g. http://localhost:8080 or the deployed URL)"),
            ("Build / version date", "2026-06-12 (Phase 4)"),
            ("UAT round", ""), ("Lead tester", ""), ("Start date", ""), ("End date", "")]
    r = 5
    for k, v in meta:
        ws.cell(row=r, column=2, value=k).font = Font(bold=True)
        ws.cell(row=r, column=3, value=v)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="How to test").font = Font(size=12, bold=True, color=brand); r += 1
    howto = [
        "1. Each tab is one functional area. Work top-to-bottom; one row = one test case.",
        "2. Follow the Steps column exactly; compare the outcome with Expected Result.",
        "2b. Sample Data / Records tells you WHICH seeded record, account or value to use — prefer it over inventing your own so results are reproducible.",
        "3. Set Status from the dropdown: Pass / Fail / Blocked / N/A.",
        "4. On Fail or Blocked, ALWAYS fill Comments (what happened, screenshots reference, error text).",
        "5. Fill Tested By + Test Date on every executed row.",
        "6. Login accounts: use the quick-login buttons on the login page, or accounts provided by the admin. Never write passwords in this file.",
        "7. Hard-refresh (Ctrl+F5) after the team deploys a fix before retesting.",
    ]
    for line in howto:
        ws.cell(row=r, column=2, value=line); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7); r += 1

    r += 1
    ws.cell(row=r, column=2, value="Status legend").font = Font(size=12, bold=True, color=brand); r += 1
    legend = [("Pass", "C8E6C9", "Works exactly as the Expected Result describes"),
              ("Fail", "FFCDD2", "Does not match the Expected Result (describe in Comments)"),
              ("Blocked", "FFE0B2", "Cannot be executed (missing data/role/dependency failed)"),
              ("N/A", "ECEFF1", "Not applicable to this round/role")]
    for name, color, desc in legend:
        c = ws.cell(row=r, column=2, value=name)
        c.fill = PatternFill("solid", fgColor=color); c.font = Font(bold=True)
        c.border = BORDER
        ws.cell(row=r, column=3, value=desc)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        r += 1

    issues = KNOWN_ISSUES.get(app_key, [])
    if issues:
        r += 1
        ws.cell(row=r, column=2, value="Known issues (do not log as new defects)").font = Font(size=12, bold=True, color="B71C1C"); r += 1
        for line in issues:
            ws.cell(row=r, column=2, value="- " + line)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 28
            r += 1

    seeds = SEED_NOTES.get(app_key, [])
    if seeds:
        r += 1
        ws.cell(row=r, column=2, value="Seeded UAT data (your starting point)").font = Font(size=12, bold=True, color=brand); r += 1
        for line in seeds:
            ws.cell(row=r, column=2, value="- " + line)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 30
            r += 1

    r += 1
    ws.cell(row=r, column=2, value="Live summary").font = Font(size=12, bold=True, color=brand); r += 1
    hdrs = ["Functional area", "Cases", "Pass", "Fail", "Blocked", "Executed %"]
    for col, h in enumerate(hdrs, start=2):
        c = ws.cell(row=r, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=brand)
        c.border = BORDER
    r += 1
    first_data = r
    for area, n in areas_counts:
        q = "'%s'" % area[:31]
        S = STATUS_COL
        ws.cell(row=r, column=2, value=area).border = BORDER
        ws.cell(row=r, column=3, value=n).border = BORDER
        ws.cell(row=r, column=4, value='=COUNTIF(%s!%s:%s,"Pass")' % (q, S, S)).border = BORDER
        ws.cell(row=r, column=5, value='=COUNTIF(%s!%s:%s,"Fail")' % (q, S, S)).border = BORDER
        ws.cell(row=r, column=6, value='=COUNTIF(%s!%s:%s,"Blocked")' % (q, S, S)).border = BORDER
        ws.cell(row=r, column=7, value='=ROUND(COUNTA(%s!%s2:%s500)/C%d*100,0)&"%%"' % (q, S, S, r)).border = BORDER
        r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = Font(bold=True)
    for col in range(3, 7):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value="=SUM(%s%d:%s%d)" % (L, first_data, L, r - 1)).font = Font(bold=True)
        ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=2).border = BORDER


# Optional CLI filter (e.g. `python uat-generator.py FL CC`) limits which apps
# are generated. Naming convention: UAT_<APP>_dd-Mon-YYYY-NN.xlsx — NN starts
# at 01 and auto-increments when a file for the same app+date already exists,
# so regenerating never overwrites a workbook testers may have filled.
import sys
from datetime import date as _date
ONLY = set(a.upper() for a in sys.argv[1:])
DATESTR = _date.today().strftime("%d-%b-%Y")

def next_out_path(out_dir, app_key):
    seq = 1
    while True:
        out = os.path.join(out_dir, "UAT_%s_%s-%02d.xlsx" % (app_key, DATESTR, seq))
        if not os.path.exists(out):
            return out
        seq += 1

total_all = 0
for app_key, app_title, areas in APPS:
    if ONLY and app_key.upper() not in ONLY:
        continue
    wb = Workbook()
    wb.remove(wb.active)
    brand = BRANDS[app_key]
    counts = []
    for area, prefix, cases in areas:
        name, n = build_area_sheet(wb, area, prefix, app_key, cases, brand)
        counts.append((name, n))
    build_instructions(wb, app_key, app_title, counts, brand)
    out_dir = os.path.join(ROOT, app_key, "UAT")
    os.makedirs(out_dir, exist_ok=True)
    out = next_out_path(out_dir, app_key)
    wb.save(out)
    n_cases = sum(n for _, n in counts)
    total_all += n_cases
    print("%-6s %3d cases, %d areas -> %s" % (app_key, n_cases, len(counts), out))
print("TOTAL  %d test cases" % total_all)
