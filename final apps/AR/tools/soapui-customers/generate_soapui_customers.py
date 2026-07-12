#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_soapui_customers.py — Excel -> SoapUI project for DoF Fusion
Receivables CreateCustomers (AR Customer integration, App 206).

Reads an Excel sheet of customers (one row per customer, headers = the wire
attribute names — see --make-template) and writes a SoapUI 5.x project XML you
can open with File > Import Project and run from your machine.

Spec: final apps/AR/docs/ws/receivables-customer-ws.md

Usage
-----
  # 1) create a fillable Excel template (all 69 columns, required marked, LOV sheet)
  python generate_soapui_customers.py --make-template customers_template.xlsx

  # 2) generate the SoapUI project from a filled sheet
  python generate_soapui_customers.py --excel customers.xlsx --env stage --out DCT-CreateCustomers.xml
  python generate_soapui_customers.py --excel customers.xlsx --env prod  --out DCT-CreateCustomers.xml

Options
-------
  --env stage|prod     gateway environment (endpoint + API key + password preset)
  --chunk N            customers per SoapUI request (default 50; each chunk
                       becomes its own request under the operation)
  --password PWD       override the Basic-auth password embedded in the project
  --apikey KEY         override the x-CentraSite-APIKey
  --endpoint URL       override the endpoint URL
  --sheet NAME         Excel sheet name (default: first sheet)
  --trx-oper OPER      default TRX_OPER when the column is empty (default CREATE)
"""
import argparse
import datetime as dt
import re
import sys
import uuid
from pathlib import Path

# ---------------------------------------------------------------------------
# Wire catalog — tag order matters (mirrors DCT_AR_WS_PKG.c_tags / the spec)
# ---------------------------------------------------------------------------
TAGS = [
    'TRX_OPER', 'CUSTOMER_TYPE', 'CUSTOMER_NAME', 'CUSTOMER_CATEGORY_NAME',
    'CUSTOMER_CLASS_NAME', 'COUNTRY_CODE', 'CITY', 'ADDRESS1', 'ADDRESS2', 'ADDRESS3',
    'ADDRESS4', 'POSTAL_CODE', 'SITE_NAME', 'SITE_CODE', 'PO_BOX_NUMBER', 'AREA',
    'SITE_END_DATE', 'PAYMENT_TERMS', 'TRADE_LICENSE_NUM', 'LEGACY_CUSTOMER_NUMBER',
    'LEGACY_CUSTOMER_SITE_NUMBER', 'CONTACT_PERSON', 'CONTACT_NUMBER', 'EMAIL_ADDRESS',
    'COLLECTOR_NAME', 'SOURCE_SYSTEM', 'PERSON_IDEN_TYPE', 'PERSON_IDENTIFIER',
    'VAT_NUMBER', 'VAT_REGISTERATION_TYPE', 'VAT_REG_START_DATE', 'VAT_REG_END_DATE',
    'TAX_REGIME_CODE', 'EMIRATE', 'ATTRIBUTE10', 'CUSTOMER_NAME_AR', 'CUST_SITE_PURPOSE',
    'ACCT_CONTACT_RESP_TYPE', 'PARENT_CUST_ACCT_NO', 'CHILD_CUST_ACCT_NO',
    'ACC_REL_ADDR_SET', 'RECIPROCAL_RELATIONSHIP', 'REL_START_DATE', 'CONTEXT_VALUE',
    'ORG_CODE', 'SITE_NUMBER', 'TAX_REG_NUM_FLAG', 'TAX_REG_NUM_REASON', 'ACCT_ESTD_DATE',
    'PREF_DELIVERY_METHOD', 'GENERATE_BILL_FLAG', 'BANK_ACCT_NAME', 'ACCT_PRIMARY_IND',
    'BANK_FROM_DATE', 'BANK_ACCT_NUMBER', 'BANK_ACCT_CURRENCY', 'BANK_NAME', 'BRANCH_NAME',
    'BANK_HOME_COUNTRY_CODE', 'BANK_ACCT_COUNTRY_CODE', 'PRIMARY_OWNER', 'IBAN',
    'SEND_STATEMENT', 'STATEMENT_CYCLE', 'SEND_CREDIT_BALANCE', 'SEND_DUNNING_LETTERS',
    'PREF_CONTACT_METHOD', 'STATEMENT_PREF_DELIVERY_METHOD', 'CUSTOMER_PROFILE',
]

REQUIRED = [
    'TRX_OPER', 'CUSTOMER_TYPE', 'CUSTOMER_NAME', 'CUSTOMER_NAME_AR',
    'CUSTOMER_CATEGORY_NAME', 'CUSTOMER_CLASS_NAME', 'COUNTRY_CODE', 'EMIRATE',
    'CITY', 'ADDRESS1', 'SITE_NAME', 'SITE_CODE', 'CUST_SITE_PURPOSE',
    'LEGACY_CUSTOMER_NUMBER', 'LEGACY_CUSTOMER_SITE_NUMBER', 'CONTACT_PERSON',
    'COLLECTOR_NAME', 'SOURCE_SYSTEM', 'VAT_REGISTERATION_TYPE',
    'TAX_REGIME_CODE', 'ORG_CODE',
]

DATE_TAGS = {'SITE_END_DATE', 'VAT_REG_START_DATE', 'VAT_REG_END_DATE',
             'ACCT_ESTD_DATE', 'REL_START_DATE', 'BANK_FROM_DATE'}

# applied when a column is missing or the cell is empty
DEFAULTS = {
    'TRX_OPER':               'CREATE',
    'CUSTOMER_TYPE':          'ORGANIZATION',
    'CUSTOMER_CATEGORY_NAME': 'High Value Customer',
    'COUNTRY_CODE':           'AE',
    'EMIRATE':                'Abu Dhabi',
    'CUST_SITE_PURPOSE':      'BOTH',
    'COLLECTOR_NAME':         'Default Collector',
    'SOURCE_SYSTEM':          'DCT_SYSTEM',
    'ORG_CODE':               'DCT',
    'VAT_REGISTERATION_TYPE': 'VAT',
    'TAX_REGIME_CODE':        'VAT_REGIME_UAE',
    'ATTRIBUTE10':            'Y',
}

LOVS = {
    'TRX_OPER':             ['CREATE', 'UPDATE-SITE-CUSTOMER'],
    'CUSTOMER_TYPE':        ['ORGANIZATION', 'PERSON'],
    'CUSTOMER_CATEGORY_NAME': ['High Value Customer'],
    'CUSTOMER_CLASS_NAME':  ['Local Company', 'International', 'Local'],
    'CUST_SITE_PURPOSE':    ['BOTH', 'BILL_TO', 'SHIP_TO'],
    'PERSON_IDEN_TYPE':     ['ID_NUMBER', 'EMAIL'],
    'VAT_REGISTERATION_TYPE': ['VAT'],
    'TAX_REGIME_CODE':      ['VAT_REGIME_UAE'],
    'TAX_REG_NUM_FLAG':     ['Y', 'N'],
    'TAX_REG_NUM_REASON':   ['NO TRADE LICENSE'],
    'PAYMENT_TERMS':        ['IMMEDIATE'],
    'EMIRATE':              ['Abu Dhabi', 'Dubai', 'Sharjah', 'Ajman',
                             'Umm Al Quwain', 'Ras Al Khaimah', 'Fujairah'],
    'COLLECTOR_NAME':       ['Default Collector'],
    'CUSTOMER_PROFILE':     ['DEFAULT'],
    'ACCT_CONTACT_RESP_TYPE': ['Statement'],
    'PREF_CONTACT_METHOD':  ['EMAIL', 'PRINT'],
    'PREF_DELIVERY_METHOD': ['EMAIL', 'PRINT'],
    'STATEMENT_PREF_DELIVERY_METHOD': ['EMAIL', 'PRINT'],
    'STATEMENT_CYCLE':      ['Monthly'],
    'GENERATE_BILL_FLAG':   ['Y', 'N'],
    'SEND_STATEMENT':       ['Y', 'N'],
    'SEND_CREDIT_BALANCE':  ['Y', 'N'],
    'SEND_DUNNING_LETTERS': ['Y', 'N'],
    'ACCT_PRIMARY_IND':     ['Y', 'N'],
}

ENVS = {
    'stage': {
        'endpoint': 'https://stage-api.abudhabi.ae/ws/Receivables_TEST/1.0',
        'apikey':   '1d9b23f1-bfca-43fc-8099-944fd7faf664',
        'password': 'P2rcpqti9n@142$',
        'wsdl':     'https://stage-api.abudhabi.ae/ws/Receivables_TEST/1.0?wsdl',
        'service':  'Receivables_TEST',
    },
    'prod': {
        'endpoint': 'https://api.abudhabi.ae/ws/Receivables/1.0',
        'apikey':   '030140d2-85be-4830-9cf2-b435b570283b',
        'password': '5fhQjevgW$aY34w',
        'wsdl':     'https://api.abudhabi.ae/ws/Receivables/1.0?wsdl',
        'service':  'Receivables',
    },
}

USERNAME = 'ADGE_DCT_INTG_SVC'
ADG_NS = 'http://xmlns.oracle.com/ADGGenericSoapWrapperApplication/ADGGenericSoapWrapperProcess'


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def xesc(v):
    return (v.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def cell_to_str(val, tag):
    """Excel cell -> wire string (dates -> DD-MM-YYYY, ints without .0)."""
    if val is None:
        return ''
    if isinstance(val, dt.datetime):
        return val.strftime('%d-%m-%Y')
    if isinstance(val, dt.date):
        return val.strftime('%d-%m-%Y')
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    s = str(val).strip()
    if tag in DATE_TAGS and re.match(r'^\d{4}-\d{2}-\d{2}', s):        # ISO -> gateway
        s = dt.datetime.strptime(s[:10], '%Y-%m-%d').strftime('%d-%m-%Y')
    return s


def norm_header(h):
    return re.sub(r'[^A-Z0-9]+', '_', str(h).strip().upper()).strip('_')


# ---------------------------------------------------------------------------
# Excel template
# ---------------------------------------------------------------------------
def make_template(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Customers'

    req_fill = PatternFill('solid', fgColor='FDE9D9')
    hdr_font = Font(bold=True)

    example = {
        'TRX_OPER': 'CREATE', 'CUSTOMER_TYPE': 'ORGANIZATION',
        'CUSTOMER_NAME': 'Example Trading LLC',
        'CUSTOMER_NAME_AR': 'شركة المثال للتجارة',
        'CUSTOMER_CATEGORY_NAME': 'High Value Customer',
        'CUSTOMER_CLASS_NAME': 'Local Company',
        'COUNTRY_CODE': 'AE', 'EMIRATE': 'Abu Dhabi', 'CITY': 'Abu Dhabi',
        'ADDRESS1': 'Corniche Road', 'POSTAL_CODE': 'NA',
        'SITE_NAME': 'Abu Dhabi', 'SITE_CODE': 'CN-1234567',
        'TRADE_LICENSE_NUM': 'CN-1234567',
        'LEGACY_CUSTOMER_NUMBER': 'CN-1234567',
        'LEGACY_CUSTOMER_SITE_NUMBER': 'CN-1234567',
        'CONTACT_PERSON': 'John Smith', 'CONTACT_NUMBER': '+971501234567',
        'EMAIL_ADDRESS': 'john@example.ae',
        'COLLECTOR_NAME': 'Default Collector', 'SOURCE_SYSTEM': 'DCT_SYSTEM',
        'VAT_NUMBER': '100123456700003',
        'VAT_REGISTERATION_TYPE': 'VAT', 'TAX_REGIME_CODE': 'VAT_REGIME_UAE',
        'ATTRIBUTE10': 'Y', 'CUST_SITE_PURPOSE': 'BOTH', 'ORG_CODE': 'DCT',
        'TAX_REG_NUM_FLAG': 'N', 'TAX_REG_NUM_REASON': 'NO TRADE LICENSE',
    }

    for col, tag in enumerate(TAGS, start=1):
        c = ws.cell(row=1, column=col, value=tag)
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center')
        notes = []
        if tag in REQUIRED:
            c.fill = req_fill
            notes.append('REQUIRED')
        if tag in DATE_TAGS:
            notes.append('Date — DD-MM-YYYY (or a real Excel date)')
        if tag in LOVS:
            notes.append('Values: ' + ' | '.join(LOVS[tag]))
        if tag in DEFAULTS:
            notes.append('Default when empty: ' + DEFAULTS[tag])
        if notes:
            c.comment = Comment('\n'.join(notes), 'AR Customer generator')
        ws.cell(row=2, column=col, value=example.get(tag, ''))
        ws.column_dimensions[get_column_letter(col)].width = max(14, min(len(tag) + 4, 32))

    ws.freeze_panes = 'A2'

    lv = wb.create_sheet('LOVs')
    lv.cell(row=1, column=1, value='Field').font = hdr_font
    lv.cell(row=1, column=2, value='Allowed values (per DoF samples — orange headers on sheet 1 are REQUIRED)').font = hdr_font
    for i, (tag, vals) in enumerate(sorted(LOVS.items()), start=2):
        lv.cell(row=i, column=1, value=tag)
        lv.cell(row=i, column=2, value=' | '.join(vals))
    lv.column_dimensions['A'].width = 34
    lv.column_dimensions['B'].width = 90

    wb.save(path)
    print(f'Template written: {path}')
    print(f'  {len(TAGS)} columns ({len(REQUIRED)} required, highlighted orange, hover a header for hints)')
    print('  Row 2 holds an EXAMPLE row — replace it with real data.')


# ---------------------------------------------------------------------------
# Excel -> customer dicts
# ---------------------------------------------------------------------------
def read_excel(path, sheet, default_oper):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    headers = {}
    for col, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        h = norm_header(cell.value)
        if h in TAGS:
            headers[col] = h
    if 'CUSTOMER_NAME' not in headers.values():
        sys.exit('ERROR: no CUSTOMER_NAME column found on the first row of the sheet. '
                 'Headers must be the wire attribute names (see --make-template).')

    customers, problems = [], []
    for rix, row in enumerate(ws.iter_rows(min_row=2), start=2):
        rec = {}
        for cell in row:
            tag = headers.get(cell.column)
            if not tag:
                continue
            v = cell_to_str(cell.value, tag)
            if v != '':
                rec[tag] = v
        if not rec:
            continue                                   # blank line
        if rec.get('CUSTOMER_NAME', '') == 'Example Trading LLC':
            continue                                   # untouched template example row
        for tag, dflt in DEFAULTS.items():
            rec.setdefault(tag, dflt)
        rec.setdefault('TRX_OPER', default_oper)
        rec.setdefault('CUSTOMER_NAME_AR', rec.get('CUSTOMER_NAME', ''))
        missing = [t for t in REQUIRED if not rec.get(t)]
        if missing:
            problems.append(f'  row {rix} ({rec.get("CUSTOMER_NAME", "?")}): missing {", ".join(missing)}')
        customers.append(rec)

    if problems:
        sys.exit('ERROR: required fields missing —\n' + '\n'.join(problems))
    if not customers:
        sys.exit('ERROR: no data rows found (the template example row is skipped automatically).')
    return customers


# ---------------------------------------------------------------------------
# envelope + SoapUI project
# ---------------------------------------------------------------------------
def build_envelope(customers):
    parts = ['<soapenv:Envelope xmlns:adg="%s" xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/">' % ADG_NS,
             '   <soapenv:Header/>',
             '   <soapenv:Body>',
             '      <adg:ADGRequestPayload>',
             '         <adg:OperationName>CreateCustomers</adg:OperationName>',
             '         <adg:InputPayload>',
             '            <CreateCustomers>',
             '               <arg0>']
    for rec in customers:
        parts.append('                  <postCustomerData>')
        for tag in TAGS:
            if tag in rec:
                parts.append('                     <%s>%s</%s>' % (tag, xesc(rec[tag]), tag))
        parts.append('                  </postCustomerData>')
    parts += ['               </arg0>',
              '            </CreateCustomers>',
              '         </adg:InputPayload>',
              '      </adg:ADGRequestPayload>',
              '   </soapenv:Body>',
              '</soapenv:Envelope>']
    xml = '\n'.join(parts)
    if ']]>' in xml:
        xml = xml.replace(']]>', ']]&gt;')             # CDATA safety
    return xml


WSDL_CONTENT = ('<wsdl:definitions name="ADGGenericSoapWrapperService" targetNamespace="{ns}" '
    'xmlns:wsdl="http://schemas.xmlsoap.org/wsdl/" xmlns:client="{ns}" '
    'xmlns:soap="http://schemas.xmlsoap.org/wsdl/soap/">'
    '<wsdl:types><schema attributeFormDefault="unqualified" elementFormDefault="qualified" '
    'targetNamespace="{ns}" xmlns="http://www.w3.org/2001/XMLSchema">'
    '<element name="ADGRequestPayload"><complexType><sequence>'
    '<element name="OperationName" type="string"/><element name="InputPayload" type="anyType"/>'
    '</sequence></complexType></element>'
    '<element name="ADGResponsePayload"><complexType><sequence>'
    '<element name="OutputPayload" type="anyType"/></sequence></complexType></element>'
    '</schema></wsdl:types>'
    '<wsdl:message name="ADGRequestMessage"><wsdl:part name="payload" element="client:ADGRequestPayload"></wsdl:part></wsdl:message>'
    '<wsdl:message name="ADGResponseMessage"><wsdl:part name="payload" element="client:ADGResponsePayload"></wsdl:part></wsdl:message>'
    '<wsdl:portType name="ADGGenericSoapWrapperProcess_ReceiveRequest_REQUEST">'
    '<wsdl:operation name="ADGSoapWraperProcess">'
    '<wsdl:input message="client:ADGRequestMessage"></wsdl:input>'
    '<wsdl:output message="client:ADGResponseMessage"></wsdl:output>'
    '</wsdl:operation></wsdl:portType>'
    '<wsdl:binding name="ADGGenericSoapWrapperProcess_ReceiveRequest_REQUEST_binding" '
    'type="client:ADGGenericSoapWrapperProcess_ReceiveRequest_REQUEST">'
    '<soap:binding style="document" transport="http://schemas.xmlsoap.org/soap/http"/>'
    '<wsdl:operation name="ADGSoapWraperProcess"><soap:operation soapAction="ADGSoapWraperProcess"/>'
    '<wsdl:input><soap:body parts="payload" use="literal"/></wsdl:input>'
    '<wsdl:output><soap:body parts="payload" use="literal"/></wsdl:output>'
    '</wsdl:operation></wsdl:binding>'
    '<wsdl:service name="{service}">'
    '<wsdl:port name="ADGGenericSoapWrapperProcess_ReceiveRequest_REQUEST_pt" '
    'binding="client:ADGGenericSoapWrapperProcess_ReceiveRequest_REQUEST_binding">'
    '<soap:address location="{endpoint}"/></wsdl:port></wsdl:service>'
    '</wsdl:definitions>')


def build_project(chunks, env_name, endpoint, apikey, password, wsdl_url, service):
    ns = ADG_NS
    header_setting = xesc('<entry key="x-CentraSite-APIKey" value="%s" xmlns="http://eviware.com/soapui/config"/>' % apikey)
    wsdl_xml = WSDL_CONTENT.format(ns=ns, service=service, endpoint=endpoint)

    calls = []
    for i, (label, envelope) in enumerate(chunks, start=1):
        calls.append(
            '<con:call id="%s" name="%s">'
            '<con:settings><con:setting id="com.eviware.soapui.impl.wsdl.WsdlRequest@request-headers">%s</con:setting></con:settings>'
            '<con:encoding>UTF-8</con:encoding>'
            '<con:endpoint>%s</con:endpoint>'
            '<con:request><![CDATA[%s]]></con:request>'
            '<con:credentials><con:username>%s</con:username><con:password>%s</con:password>'
            '<con:selectedAuthProfile>Basic</con:selectedAuthProfile>'
            '<con:addedBasicAuthenticationTypes>Basic</con:addedBasicAuthenticationTypes>'
            '<con:authType>Global HTTP Settings</con:authType></con:credentials>'
            '<con:jmsConfig JMSDeliveryMode="PERSISTENT"/><con:jmsPropertyConfig/>'
            '<con:wsaConfig mustUnderstand="NONE" version="200508" action="ADGSoapWraperProcess"/>'
            '<con:wsrmConfig version="1.2"/></con:call>'
            % (uuid.uuid4(), xesc(label), header_setting, endpoint,
               envelope, USERNAME, xesc(password)))

    project = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<con:soapui-project id="%(pid)s" activeEnvironment="Default" name="%(pname)s" '
        'resourceRoot="" soapui-version="5.9.1" abortOnError="false" runType="SEQUENTIAL" '
        'xmlns:con="http://eviware.com/soapui/config"><con:settings/>'
        '<con:interface xsi:type="con:WsdlInterface" id="%(iid)s" wsaVersion="NONE" '
        'name="ADGGenericSoapWrapperProcess_ReceiveRequest_REQUEST_binding" type="wsdl" '
        'bindingName="{%(ns)s}ADGGenericSoapWrapperProcess_ReceiveRequest_REQUEST_binding" '
        'soapVersion="1_1" anonymous="optional" definition="%(wsdl)s" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><con:settings/>'
        '<con:definitionCache type="TEXT" rootPart="%(wsdl)s">'
        '<con:part><con:url>%(wsdl)s</con:url><con:content><![CDATA[%(wsdlxml)s]]></con:content>'
        '<con:type>http://schemas.xmlsoap.org/wsdl/</con:type></con:part></con:definitionCache>'
        '<con:endpoints><con:endpoint>%(endpoint)s</con:endpoint></con:endpoints>'
        '<con:operation id="%(oid)s" isOneWay="false" action="ADGSoapWraperProcess" '
        'name="ADGSoapWraperProcess" bindingOperationName="ADGSoapWraperProcess" '
        'type="Request-Response" inputName="" receivesAttachments="false" '
        'sendsAttachments="false" anonymous="optional"><con:settings/>%(calls)s</con:operation>'
        '</con:interface><con:properties/><con:wssContainer/>'
        '<con:oAuth2ProfileContainer/><con:oAuth1ProfileContainer/>'
        '<con:sensitiveInformation/></con:soapui-project>\n'
    ) % {
        'pid': uuid.uuid4(), 'iid': uuid.uuid4(), 'oid': uuid.uuid4(),
        'pname': 'DCT-CreateCustomers-%s' % env_name.upper(),
        'ns': ns, 'wsdl': wsdl_url, 'wsdlxml': wsdl_xml,
        'endpoint': endpoint, 'calls': ''.join(calls),
    }
    return project


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--make-template', metavar='OUT.xlsx', help='write a fillable Excel template and exit')
    ap.add_argument('--excel', help='input Excel sheet with customer data')
    ap.add_argument('--sheet', help='sheet name (default: first sheet)')
    ap.add_argument('--env', choices=['stage', 'prod'], default='stage')
    ap.add_argument('--out', help='output SoapUI project XML (default: DCT-CreateCustomers-<env>-soapui-project.xml)')
    ap.add_argument('--chunk', type=int, default=50, help='customers per request (default 50)')
    ap.add_argument('--password', help='override the Basic-auth password')
    ap.add_argument('--apikey', help='override the x-CentraSite-APIKey')
    ap.add_argument('--endpoint', help='override the endpoint URL')
    ap.add_argument('--trx-oper', default='CREATE', choices=['CREATE', 'UPDATE-SITE-CUSTOMER'],
                    help='default TRX_OPER when the column is empty')
    args = ap.parse_args()

    if args.make_template:
        make_template(args.make_template)
        return
    if not args.excel:
        ap.error('provide --excel IN.xlsx (or --make-template OUT.xlsx)')

    env = ENVS[args.env]
    endpoint = args.endpoint or env['endpoint']
    apikey   = args.apikey   or env['apikey']
    password = args.password or env['password']
    out      = Path(args.out or ('DCT-CreateCustomers-%s-soapui-project.xml' % args.env))

    customers = read_excel(args.excel, args.sheet, args.trx_oper)
    chunks = []
    for i in range(0, len(customers), args.chunk):
        batch = customers[i:i + args.chunk]
        label = ('CreateCustomers %d-%d of %d' % (i + 1, i + len(batch), len(customers))
                 if len(customers) > args.chunk else 'CreateCustomers (%d customers)' % len(batch))
        chunks.append((label, build_envelope(batch)))

    out.write_text(build_project(chunks, args.env, endpoint, apikey, password,
                                 env['wsdl'], env['service']), encoding='utf-8')

    print('SoapUI project written: %s' % out)
    print('  env=%s  endpoint=%s' % (args.env, endpoint))
    print('  customers=%d  requests=%d (chunk=%d)' % (len(customers), len(chunks), args.chunk))
    print('Open in SoapUI: File > Import Project, expand ADGSoapWraperProcess, run each request.')


if __name__ == '__main__':
    main()
